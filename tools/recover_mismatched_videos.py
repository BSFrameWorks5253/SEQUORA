#!/usr/bin/env python3
"""
===============================================================================
 SEQUORA — Jumbled Video Files Recovery & Reshuffler Utility
===============================================================================
 Automatically scans Photo Data subfolders (e.g., 01-Q, 16-Q, 17-Q, 02-MZ, 03-Z,
 04-BZ, 05-BQ, 06-BM, 07-M, etc.) within each Date folder, detects misplaced
 video files (where video sequence != folder sequence), and safely moves them
 into their correct destination subfolder inside the SAME DATE FOLDER.

 Features:
   - Preserves strict date-folder isolation (only shuffles within the same date).
   - Parses camera clip prefixes (001, C0001, CAM1) and sequence numbers (01, 16, 17).
   - Moves associated _v.jpg / _v.png video thumbnails alongside video files.
   - Full Dry-Run preview before any files are moved.
   - Generates formatted Excel (.xlsx) & CSV Recovery Manifests in Document/ folder.
   - Dual interface: Easy Graphical UI (Tkinter) and CLI for automation.
===============================================================================
"""

import os
import sys
import re
import csv
import time
import shutil
import argparse
from pathlib import Path

VIDEO_EXTS = {".mp4", ".mov", ".mkv", ".avi", ".m4v", ".wmv", ".flv", ".mts", ".m2ts"}
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}


def normalize_key(s: str) -> str:
    return re.sub(r'[\s_\-]+', ' ', s.lower()).strip()


def parse_sequence_name(raw_name: str) -> dict:
    ext = os.path.splitext(raw_name)[1]
    name_without_ext = os.path.splitext(raw_name)[0].strip() if ext else raw_name.strip()

    # 1. Strip trailing video/thumbnail suffixes (-its-v, _v, -v, _p, -p)
    clean = re.sub(r'[-_](?:its|v|p)+(?:[-_]v|[-_]p)?$', '', name_without_ext, flags=re.I).strip()
    # 2. Strip leading date format if present (e.g. 2026-09-01_ or 1448-03-12_)
    clean = re.sub(r'^\d{4}[-_.]\d{2}[-_.]\d{2}[\s_-]*', '', clean).strip()

    # 3. Check for camera clip prefix like '001 01-Q', '002 16-Q', 'C0001_01-Q', 'CAM1_01-Q', '01_01-Q'
    m_cam = re.match(r'^(?:[a-zA-Z]{0,4}\d{2,5})[\s_\-]+(\d+[a-zA-Z]?[\s_\-]+[a-zA-Z0-9]+(?:[\s_\-].*)?)$', clean)
    if m_cam:
        clean = m_cam.group(1).strip()

    # 4. Extract sequence number (e.g. '01', '16', '17', '01a') and code part (e.g. 'Q', 'MZ', 'Z', 'BZ', 'BQ', 'BM', 'M', 'KG', 'N')
    m_seq = re.match(r'^(\d+[a-zA-Z]?)(?:[\s_\-]+([a-zA-Z0-9]+))?', clean)
    sequence = m_seq.group(1) if m_seq else ""
    code_part = m_seq.group(2) if (m_seq and m_seq.group(2)) else ""

    norm_seq = f"{int(sequence):02d}" if sequence.isdigit() else sequence.lower()
    norm_code = code_part.lower()
    token = f"{norm_seq}-{norm_code}" if (norm_seq and norm_code) else (norm_seq or clean.lower())
    prefix = code_part.upper() if code_part else ""

    return {
        "raw": raw_name,
        "cleanName": clean,
        "sequence": norm_seq,
        "namePart": code_part or clean,
        "prefix": prefix,
        "code": norm_code,
        "token": token
    }


def is_match_pair(vid_parsed: dict, pf_parsed: dict) -> bool:
    v_seq = vid_parsed.get("sequence", "")
    p_seq = pf_parsed.get("sequence", "")
    v_code = vid_parsed.get("code", "")
    p_code = pf_parsed.get("code", "")
    v_tok = vid_parsed.get("token", "")
    p_tok = pf_parsed.get("token", "")
    v_clean = vid_parsed.get("cleanName", "").lower()
    p_clean = pf_parsed.get("cleanName", "").lower()

    if v_seq and p_seq and v_seq != p_seq:
        return False

    if v_clean and p_clean and v_clean == p_clean:
        return True

    if v_tok and p_tok and v_tok == p_tok:
        return True

    if v_seq and p_seq and v_seq == p_seq:
        if v_code and p_code and v_code == p_code:
            return True
        if not v_code or not p_code:
            return True

    v_norm = re.sub(r'[\s_\-]+', ' ', v_clean).strip()
    p_norm = re.sub(r'[\s_\-]+', ' ', p_clean).strip()
    if v_norm and p_norm and v_norm == p_norm:
        return True

    return False


def is_date_folder(name: str) -> bool:
    if re.search(r'\d{4}[-_.]\d{2}[-_.]\d{2}', name) or re.search(r'\d{2}[-_.]\d{2}[-_.]\d{4}', name):
        return True
    if re.match(r'^(?:day|date|event|session)[\s_\-]*\d+', name, re.I):
        return True
    return False


def find_associated_video_thumbnail(folder_path: str, video_filename: str):
    vid_stem = os.path.splitext(video_filename)[0].lower()
    vid_base = re.sub(r'[-_][vp]$', '', vid_stem)
    vid_parsed = parse_sequence_name(video_filename)
    try:
        for f in os.listdir(folder_path):
            fp = os.path.join(folder_path, f)
            if not os.path.isfile(fp):
                continue
            ext = os.path.splitext(f)[1].lower()
            if ext in IMAGE_EXTS:
                f_lower = f.lower()
                f_stem = os.path.splitext(f)[0].lower()
                f_parsed = parse_sequence_name(f)
                if (f_lower.endswith('_v.jpg') or f_lower.endswith('_v.png') or f_lower.endswith('_v.jpeg')) and \
                   (f_parsed['token'] == vid_parsed['token'] or f_stem.startswith(vid_base)):
                    return fp
    except Exception:
        pass
    return None


def scan_for_mismatches(photo_root_dir: str) -> list:
    photo_root = os.path.abspath(photo_root_dir)
    if not os.path.isdir(photo_root):
        raise ValueError(f"Photo directory does not exist: {photo_root}")

    # Discover date folders
    date_folders = []
    for entry in os.scandir(photo_root):
        if entry.is_dir() and is_date_folder(entry.name):
            date_folders.append(entry)

    # If no date folders at top level, treat photo_root as single date folder
    if not date_folders:
        class FakeEntry:
            def __init__(self, p):
                self.path = p
                self.name = os.path.basename(p)
                self.is_dir = lambda: True
        date_folders = [FakeEntry(photo_root)]

    mismatches = []

    for date_entry in date_folders:
        date_folder_path = date_entry.path
        date_folder_name = date_entry.name

        # Discover all subfolders inside this date folder
        subfolders = []
        for item in os.scandir(date_folder_path):
            if item.is_dir():
                subfolders.append({
                    "name": item.name,
                    "path": item.path,
                    "parsed": parse_sequence_name(item.name)
                })

        # Scan each subfolder for video files
        for sf in subfolders:
            curr_folder_path = sf["path"]
            curr_folder_name = sf["name"]
            curr_folder_parsed = sf["parsed"]

            try:
                for fname in os.listdir(curr_folder_path):
                    fpath = os.path.join(curr_folder_path, fname)
                    if not os.path.isfile(fpath):
                        continue
                    ext = os.path.splitext(fname)[1].lower()
                    if ext not in VIDEO_EXTS:
                        continue

                    vid_parsed = parse_sequence_name(fname)

                    # Check if this video matches current folder
                    if is_match_pair(vid_parsed, curr_folder_parsed):
                        # Video is in the CORRECT folder!
                        continue

                    # MISMATCH DETECTED: Video belongs to a different subfolder under SAME date!
                    # Find matching sibling folder under same date folder
                    correct_sf = None
                    for target_sf in subfolders:
                        if is_match_pair(vid_parsed, target_sf["parsed"]):
                            correct_sf = target_sf
                            break

                    target_folder_path = ""
                    target_folder_name = ""
                    if correct_sf:
                        target_folder_path = correct_sf["path"]
                        target_folder_name = correct_sf["name"]
                    else:
                        ideal_name = f"{vid_parsed['sequence']}-{vid_parsed['prefix']}" if (vid_parsed['sequence'] and vid_parsed['prefix']) else vid_parsed['cleanName']
                        target_folder_path = os.path.join(date_folder_path, ideal_name)
                        target_folder_name = ideal_name

                    thumb_path = find_associated_video_thumbnail(curr_folder_path, fname)
                    thumb_target_path = os.path.join(target_folder_path, os.path.basename(thumb_path)) if thumb_path else ""

                    mismatches.append({
                        "dateFolderName": date_folder_name,
                        "dateFolderPath": date_folder_path,
                        "videoFilename": fname,
                        "videoSourcePath": fpath,
                        "videoParsed": vid_parsed,
                        "currentFolderName": curr_folder_name,
                        "currentFolderPath": curr_folder_path,
                        "targetFolderName": target_folder_name,
                        "targetFolderPath": target_folder_path,
                        "targetVideoPath": os.path.join(target_folder_path, fname),
                        "thumbSourcePath": thumb_path or "",
                        "thumbTargetPath": thumb_target_path,
                        "fileSizeBytes": os.path.getsize(fpath) if os.path.exists(fpath) else 0
                    })
            except Exception as e:
                print(f"Error scanning {curr_folder_path}: {e}")

    return mismatches


def execute_recovery(mismatches: list, photo_root_dir: str, dry_run: bool = False) -> dict:
    recovered = []
    errors = []

    doc_dir = os.path.join(photo_root_dir, "Document")
    os.makedirs(doc_dir, exist_ok=True)
    ts = time.strftime("%Y%m%d_%H%M%S")
    date_display = time.strftime("%Y-%m-%d %H:%M:%S")

    for item in mismatches:
        src_v = item["videoSourcePath"]
        dst_v = item["targetVideoPath"]
        dst_folder = item["targetFolderPath"]
        src_t = item["thumbSourcePath"]
        dst_t = item["thumbTargetPath"]

        if dry_run:
            recovered.append({**item, "status": "PREVIEW_OK"})
            continue

        try:
            os.makedirs(dst_folder, exist_ok=True)

            # Check if destination file exists already
            if os.path.exists(dst_v) and os.path.abspath(src_v) != os.path.abspath(dst_v):
                base, ext = os.path.splitext(dst_v)
                copy_idx = 1
                while os.path.exists(dst_v):
                    dst_v = f"{base}_copy{copy_idx}{ext}"
                    copy_idx += 1
                item["targetVideoPath"] = dst_v

            if os.path.abspath(src_v) != os.path.abspath(dst_v):
                shutil.move(src_v, dst_v)

            # Move thumbnail if found
            if src_t and os.path.exists(src_t) and dst_t:
                if os.path.abspath(src_t) != os.path.abspath(dst_t):
                    if os.path.exists(dst_t):
                        os.remove(dst_t)
                    shutil.move(src_t, dst_t)

            recovered.append({**item, "status": "RECOVERED_SUCCESS"})
        except Exception as ex:
            errors.append(f"{item['videoFilename']}: {str(ex)}")
            recovered.append({**item, "status": f"ERROR: {str(ex)}"})

    # Export Excel & CSV Recovery Manifest
    xlsx_path = os.path.join(doc_dir, f"Recovery_Manifest_{ts}.xlsx")
    csv_path = os.path.join(doc_dir, f"Recovery_Manifest_{ts}.csv")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recovery Manifest"

        ws.merge_cells("A1:K1")
        ws["A1"] = "SEQUORA Studio — Jumbled Video Recovery & Reshuffle Manifest"
        ws["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        meta_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
        meta_font = Font(name="Segoe UI", size=10, bold=True, color="333333")
        val_font = Font(name="Segoe UI", size=10, color="222222")

        ws["A2"] = "Execution Date:"
        ws["A2"].font = meta_font
        ws["A2"].fill = meta_fill
        ws["B2"] = date_display
        ws["B2"].font = val_font

        ws["D2"] = "Total Reshuffled:"
        ws["D2"].font = meta_font
        ws["D2"].fill = meta_fill
        ws["E2"] = len(recovered)
        ws["E2"].font = val_font

        ws["G2"] = "Mode:"
        ws["G2"].font = meta_font
        ws["G2"].fill = meta_fill
        ws["H2"] = "DRY RUN (PREVIEW)" if dry_run else "EXECUTED RESHUFFLE"
        ws["H2"].font = val_font

        headers = [
            "SR NO",
            "DATE FOLDER",
            "SEQUENCE / TOKEN",
            "VIDEO FILENAME",
            "FROM INCORRECT FOLDER",
            "TO CORRECT PHOTO FOLDER",
            "SOURCE PATH BEFORE",
            "TARGET PATH AFTER",
            "THUMBNAIL MOVED",
            "STATUS",
            "TIMESTAMP"
        ]

        hdr_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        hdr_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        thin_side = Side(style="thin", color="D9D9D9")
        table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        ws.row_dimensions[4].height = 24
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=hdr)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = table_border

        zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        for row_idx, r in enumerate(recovered, 5):
            is_even = (row_idx % 2 == 0)
            row_data = [
                row_idx - 4,
                r.get("dateFolderName", ""),
                r.get("videoParsed", {}).get("token", ""),
                r.get("videoFilename", ""),
                r.get("currentFolderName", ""),
                r.get("targetFolderName", ""),
                r.get("videoSourcePath", ""),
                r.get("targetVideoPath", ""),
                "YES" if r.get("thumbSourcePath") else "NO",
                r.get("status", ""),
                date_display
            ]
            ws.row_dimensions[row_idx].height = 20
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Segoe UI", size=9, color="222222")
                if is_even:
                    cell.fill = zebra_fill
                cell.border = table_border
                if col_idx in (1, 3, 5, 6, 9, 10):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in ws.columns:
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            max_len = max(len(str(c.value or "")) for c in col if c.row > 1) if col else 12
            ws.column_dimensions[col_letter].width = max(12, min(max_len + 3, 55))

        wb.save(xlsx_path)
    except Exception as ex_wb:
        print(f"Excel export notice: {ex_wb}")

    # CSV Export
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f_csv:
        writer = csv.writer(f_csv)
        writer.writerow([
            "SR NO",
            "DATE FOLDER",
            "SEQUENCE",
            "VIDEO FILENAME",
            "FROM INCORRECT FOLDER",
            "TO CORRECT PHOTO FOLDER",
            "SOURCE PATH BEFORE",
            "TARGET PATH AFTER",
            "THUMBNAIL MOVED",
            "STATUS",
            "TIMESTAMP"
        ])
        for idx, r in enumerate(recovered, 1):
            writer.writerow([
                idx,
                r.get("dateFolderName", ""),
                r.get("videoParsed", {}).get("token", ""),
                r.get("videoFilename", ""),
                r.get("currentFolderName", ""),
                r.get("targetFolderName", ""),
                r.get("videoSourcePath", ""),
                r.get("targetVideoPath", ""),
                "YES" if r.get("thumbSourcePath") else "NO",
                r.get("status", ""),
                date_display
            ])

    return {
        "total": len(mismatches),
        "successful": len(recovered) - len(errors),
        "errors": errors,
        "xlsxReport": xlsx_path if os.path.exists(xlsx_path) else "",
        "csvReport": csv_path,
        "recovered": recovered
    }


def launch_gui():
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox

    root = tk.Tk()
    root.title("SEQUORA — Jumbled Video Recovery & Reshuffler")
    root.geometry("980x640")
    root.minsize(800, 500)

    style = ttk.Style()
    try:
        style.theme_use("clam")
    except Exception:
        pass

    header_frame = tk.Frame(root, bg="#1F4E79", height=70)
    header_frame.pack(fill="x")

    title_lbl = tk.Label(header_frame, text="SEQUORA — Jumbled Video Recovery Tool", font=("Segoe UI", 16, "bold"), fg="#FFFFFF", bg="#1F4E79")
    title_lbl.pack(anchor="w", padx=20, pady=(12, 2))

    sub_lbl = tk.Label(header_frame, text="Re-shuffle misplaced video clips (001 01-Q, 16-Q, 02-MZ, 03-Z...) to their matching subfolder within the same date folder.", font=("Segoe UI", 9), fg="#DCE6F1", bg="#1F4E79")
    sub_lbl.pack(anchor="w", padx=20, pady=(0, 10))

    ctrl_frame = tk.Frame(root, padx=16, pady=12)
    ctrl_frame.pack(fill="x")

    tk.Label(ctrl_frame, text="Photo Data Directory:", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w", pady=4)
    path_var = tk.StringVar()
    path_entry = ttk.Entry(ctrl_frame, textvariable=path_var, width=70)
    path_entry.grid(row=0, column=1, sticky="ew", padx=8, pady=4)
    ctrl_frame.columnconfigure(1, weight=1)

    def browse_dir():
        d = filedialog.askdirectory(title="Select Photo Data Directory")
        if d:
            path_var.set(d)
            do_scan()

    browse_btn = ttk.Button(ctrl_frame, text="Browse...", command=browse_dir)
    browse_btn.grid(row=0, column=2, padx=4, pady=4)

    status_var = tk.StringVar(value="Select Photo directory to scan for misplaced videos.")
    status_lbl = tk.Label(root, textvariable=status_var, font=("Segoe UI", 9, "italic"), fg="#555555", anchor="w", padx=20)
    status_lbl.pack(fill="x")

    table_frame = tk.Frame(root, padx=16, pady=6)
    table_frame.pack(fill="both", expand=True)

    cols = ("Date Folder", "Video Filename", "Current Folder (Wrong)", "Correct Target Folder", "Token", "Size")
    tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="extended")

    tree.heading("Date Folder", text="Date Folder")
    tree.heading("Video Filename", text="Video Filename")
    tree.heading("Current Folder (Wrong)", text="Current Folder (Wrong)")
    tree.heading("Correct Target Folder", text="Correct Target Folder")
    tree.heading("Token", text="Token")
    tree.heading("Size", text="Size")

    tree.column("Date Folder", width=120)
    tree.column("Video Filename", width=200)
    tree.column("Current Folder (Wrong)", width=170)
    tree.column("Correct Target Folder", width=170)
    tree.column("Token", width=80, anchor="center")
    tree.column("Size", width=80, anchor="e")

    v_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=v_scroll.set)
    tree.pack(side="left", fill="both", expand=True)
    v_scroll.pack(side="right", fill="y")

    current_mismatches = []

    def do_scan():
        d = path_var.get().strip()
        if not d or not os.path.isdir(d):
            messagebox.showwarning("Warning", "Please select a valid Photo directory first.")
            return

        tree.delete(*tree.get_children())
        status_var.set("Scanning for misplaced video files...")
        root.update_idletasks()

        try:
            mismatches = scan_for_mismatches(d)
            current_mismatches.clear()
            current_mismatches.extend(mismatches)

            if not mismatches:
                status_var.set("✅ Perfect! No misplaced video files found. All videos match their folders.")
                messagebox.showinfo("Scan Complete", "No misplaced video files found! All videos are in their proper matching subfolders.")
                return

            for item in mismatches:
                sz_mb = f"{item['fileSizeBytes'] / (1024*1024):.1f} MB"
                tree.insert("", "end", values=(
                    item["dateFolderName"],
                    item["videoFilename"],
                    f"⚠️ {item['currentFolderName']}",
                    f"👉 {item['targetFolderName']}",
                    item["videoParsed"]["token"],
                    sz_mb
                ))

            status_var.set(f"⚠️ Found {len(mismatches)} misplaced video files ready to be reshuffled into their correct subfolders.")
        except Exception as e:
            status_var.set(f"Error during scan: {e}")
            messagebox.showerror("Error", str(e))

    def do_recover():
        if not current_mismatches:
            messagebox.showinfo("Notice", "No misplaced files to recover. Please scan a directory first.")
            return

        confirm = messagebox.askyesno(
            "Confirm Reshuffle",
            f"Are you sure you want to move {len(current_mismatches)} misplaced video files into their correct matching subfolders?\n\n"
            "• Moves will occur strictly within the same Date folder.\n"
            "• Associated _v.jpg video thumbnails will also be moved.\n"
            "• An Excel & CSV undo manifest will be saved in the Document/ folder."
        )
        if not confirm:
            return

        status_var.set("Reshuffling files into correct subfolders...")
        root.update_idletasks()

        try:
            res = execute_recovery(current_mismatches, path_var.get().strip(), dry_run=False)
            status_var.set(f"✅ Successfully reshuffled {res['successful']} files! Excel manifest saved to Document folder.")
            msg = f"Recovery Complete!\n\n• Reshuffled: {res['successful']} video files\n• Errors: {len(res['errors'])}\n\nManifest saved to:\n{res['xlsxReport']}"
            messagebox.showinfo("Success", msg)
            do_scan()
        except Exception as e:
            status_var.set(f"Recovery failed: {e}")
            messagebox.showerror("Recovery Error", str(e))

    btn_frame = tk.Frame(root, padx=16, pady=12)
    btn_frame.pack(fill="x")

    scan_btn = ttk.Button(btn_frame, text="🔍 Scan for Misplaced Videos", command=do_scan)
    scan_btn.pack(side="left", padx=4)

    recover_btn = tk.Button(btn_frame, text="⚡ Re-Shuffle into Correct Subfolders", font=("Segoe UI", 10, "bold"), bg="#1F4E79", fg="#FFFFFF", padx=16, pady=6, command=do_recover)
    recover_btn.pack(side="right", padx=4)

    root.mainloop()


def main():
    parser = argparse.ArgumentParser(description="SEQUORA Jumbled Video Files Recovery & Reshuffler")
    parser.add_argument("--photo-dir", "-p", help="Path to Photo Data directory")
    parser.add_argument("--dry-run", "-d", action="store_true", help="Preview moves without moving files")
    parser.add_argument("--cli", action="store_true", help="Run in CLI mode instead of GUI")

    args = parser.parse_args()

    if not args.photo_dir and not args.cli and len(sys.argv) == 1:
        launch_gui()
        return

    if not args.photo_dir:
        print("Error: --photo-dir is required when running in CLI mode.")
        sys.exit(1)

    photo_dir = os.path.abspath(args.photo_dir)
    print(f"Scanning for misplaced video files in: {photo_dir}")
    mismatches = scan_for_mismatches(photo_dir)
    print(f"Found {len(mismatches)} misplaced video file(s).")

    if not mismatches:
        print("All video files are already in their correct matching subfolders!")
        return

    for idx, item in enumerate(mismatches, 1):
        print(f"[{idx:3d}] Date: {item['dateFolderName']} | Video: {item['videoFilename']} | Wrong: {item['currentFolderName']} -> Correct: {item['targetFolderName']}")

    if args.dry_run:
        print("\n[DRY RUN] No files were moved. Generating preview manifest...")
    else:
        print(f"\nExecuting reshuffle for {len(mismatches)} files...")

    res = execute_recovery(mismatches, photo_dir, dry_run=args.dry_run)
    print(f"\nDone! Successful: {res['successful']} | Errors: {len(res['errors'])}")
    if res["xlsxReport"]:
        print(f"Excel Manifest: {res['xlsxReport']}")
    if res["csvReport"]:
        print(f"CSV Manifest: {res['csvReport']}")


if __name__ == "__main__":
    main()
