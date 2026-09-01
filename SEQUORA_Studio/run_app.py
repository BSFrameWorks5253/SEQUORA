#!/usr/bin/env python3
"""
===============================================================================
 One Tap Studio — Native Qt6/QML Backend Runtime
===============================================================================
 Connects QML GUI directly to high-performance Python engines for:
  1. Photo Matcher (_U and _R renaming engine with safety validation)
  2. Video Transfer (Smart date & sequence matcher with audio/preview collision)
  3. Thumbnail Separator (Recursive _P and _V thumbnail extraction & undo)
===============================================================================
"""

import os
import sys
import re
import csv
import json
import time
import shutil
import threading
from pathlib import Path

# Add PySide6 DLL directory for Windows Qt plugin resolution
try:
    import PySide6
    pyside6_dir = os.path.dirname(PySide6.__file__)
    if hasattr(os, "add_dll_directory") and os.path.isdir(pyside6_dir):
        os.add_dll_directory(pyside6_dir)
    os.environ["PATH"] = pyside6_dir + os.pathsep + os.environ.get("PATH", "")
except Exception:
    pass

# Add project root to sys.path
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

# Set Windows taskbar App ID globally before creating any UI or COM components
APP_USER_MODEL_ID = "BSFrameWorks.SEQUORA.Studio.v3"
if sys.platform == "win32":
    try:
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(APP_USER_MODEL_ID)
    except Exception:
        pass

from PySide6.QtCore import QObject, Signal, Slot, Property, QUrl, Qt
from PySide6.QtWidgets import QApplication, QFileDialog
from PySide6.QtGui import QIcon, QFont
from PySide6.QtQml import QQmlApplicationEngine
from PySide6.QtNetwork import QLocalServer, QLocalSocket


# Configure QtWebEngine Chromium Flags
os.environ["QTWEBENGINE_CHROMIUM_FLAGS"] = (
    "--enable-gpu-rasterization --no-default-browser-check"
)

try:
    from PySide6.QtWebEngineQuick import QtWebEngineQuick
    QtWebEngineQuick.initialize()
except Exception as e:
    print(f"Notice: QtWebEngineQuick initialization: {e}")


# Import core business logic from app services if available, else local fallback
try:
    from app.services.directory_scanner import DirectoryScanner
    from app.services.matcher import PhotoMatcher
    from app.services.rename_engine import RenameEngine
    from app.services.report_generator import ReportGenerator
    from app.models.config import AppConfig
except ImportError:
    DirectoryScanner = None
    PhotoMatcher = None
    RenameEngine = None
    ReportGenerator = None
    AppConfig = None


def to_py_variant(val):
    if hasattr(val, "toVariant"):
        return val.toVariant()
    return val


class PhotoMatcherEngineWrapper(QObject):
    scanningChanged = Signal()
    renamingChanged = Signal()
    configChanged = Signal()
    scanCompleted = Signal("QVariant")
    scanError = Signal(str)
    preStatsReady = Signal("QVariant")
    renameCompleted = Signal("QVariant", "QVariant", "QVariant", "QVariant")
    reportExported = Signal(str, str)
    error = Signal(str)

    PHOTO_EXTS = {'.cr2', '.cr3', '.nef', '.arw', '.raf', '.dng', '.orf', '.rw2', '.pef', '.srw', '.heic', '.jpg', '.jpeg', '.png'}

    def __init__(self, parent=None):
        super().__init__(parent)
        self._scanning = False
        self._renaming = False
        self._config_file = SCRIPT_DIR / "app_config.json"
        self._config = {
            "theme": "dark",
            "zoomLevel": 1.0,
            "dryRun": True,
            "backupEnabled": True,
            "autoExportReport": True
        }
        self.load_config()
        self._last_scan_result = None
        self._last_path = ""

    def load_config(self):
        if self._config_file.exists():
            try:
                with open(self._config_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    if isinstance(data, dict):
                        self._config.update(data)
            except Exception as e:
                print(f"Notice: loading app config: {e}")

    def save_config_file(self):
        try:
            with open(self._config_file, "w", encoding="utf-8") as f:
                json.dump(self._config, f, indent=2)
        except Exception as e:
            print(f"Notice: saving app config: {e}")

    @Property(bool, notify=scanningChanged)
    def scanning(self):
        return self._scanning

    @Property(bool, notify=renamingChanged)
    def renaming(self):
        return self._renaming

    @Property(dict, notify=configChanged)
    def config(self):
        return self._config

    @Slot(str)
    def setConfigTheme(self, theme_name):
        self._config["theme"] = theme_name
        self.save_config_file()
        self.configChanged.emit()

    @Slot("QVariant")
    def saveConfig(self, new_config):
        new_config = to_py_variant(new_config)
        if isinstance(new_config, dict):
            self._config.update(new_config)
            self.save_config_file()
            self.configChanged.emit()

    @Slot(str)
    def getPreStats(self, dir_path):
        if not dir_path:
            return
        clean_path = os.path.abspath(os.path.normpath(dir_path.replace("file:///", "").strip()))
        if not os.path.isdir(clean_path):
            return

        def worker():
            try:
                date_folders = set()
                total_orig = 0
                total_rem = 0

                def is_date_folder_name(name):
                    if re.search(r'\d{4}[-_.]\d{2}[-_.]\d{2}', name) or re.search(r'\d{2}[-_.]\d{2}[-_.]\d{4}', name):
                        return True
                    if re.match(r'^(?:day|date|event|session)[\s_\-]*\d+', name, re.I):
                        return True
                    return False

                for entry in os.scandir(clean_path):
                    if entry.is_dir() and is_date_folder_name(entry.name):
                        date_folders.add(entry.name)

                for root, dirs, files in os.walk(clean_path):
                    rel = os.path.relpath(root, clean_path)
                    parts = [p for p in rel.replace('\\', '/').split('/') if p and p != '.']
                    is_rem = any('remaining' in p.lower() for p in parts)
                    for f in files:
                        ext = os.path.splitext(f)[1].lower()
                        if ext in self.PHOTO_EXTS:
                            stem = os.path.splitext(f)[0]
                            if is_rem or stem.endswith('_R') or stem.endswith('_U') or stem.endswith('_r') or stem.endswith('_u'):
                                total_rem += 1
                            else:
                                total_orig += 1

                stats_payload = {
                    "pre_stats": {
                        "date_folder_count": len(date_folders) if date_folders else (1 if total_rem > 0 else 0),
                        "total_original_photos": total_orig,
                        "total_remaining_photos": total_rem
                    }
                }
                self.preStatsReady.emit(stats_payload)
            except Exception as e:
                print(f"PreStats error: {e}")

        threading.Thread(target=worker, daemon=True).start()

    @Slot(str)
    def scan(self, dir_path):
        self.scanDirectory(dir_path)

    @Slot(str)
    def scanDirectory(self, dir_path):
        if not dir_path:
            self.error.emit("Please select a target directory.")
            self.scanError.emit("Please select a target directory.")
            return
        clean_path = os.path.abspath(os.path.normpath(dir_path.replace("file:///", "").strip()))
        if not os.path.isdir(clean_path):
            err = f"Directory does not exist: {clean_path}"
            self.error.emit(err)
            self.scanError.emit(err)
            return

        self._last_path = clean_path
        self._scanning = True
        self.scanningChanged.emit()

        def worker():
            try:
                date_folders = set()
                orig_photos_by_date = {}
                rem_photos_list = []

                def is_date_folder_name(name):
                    if re.search(r'\d{4}[-_.]\d{2}[-_.]\d{2}', name) or re.search(r'\d{2}[-_.]\d{2}[-_.]\d{4}', name):
                        return True
                    if re.match(r'^(?:day|date|event|session)[\s_\-]*\d+', name, re.I):
                        return True
                    return False

                for entry in os.scandir(clean_path):
                    if entry.is_dir() and is_date_folder_name(entry.name):
                        date_folders.add(entry.name)

                for root, dirs, files in os.walk(clean_path):
                    rel = os.path.relpath(root, clean_path)
                    parts = [p for p in rel.replace('\\', '/').split('/') if p and p != '.']
                    d_name = parts[0] if parts and parts[0] in date_folders else (parts[0] if parts else os.path.basename(clean_path))
                    is_rem = any('remaining' in p.lower() for p in parts)

                    for f in files:
                        ext = os.path.splitext(f)[1].lower()
                        if ext in self.PHOTO_EXTS:
                            stem = os.path.splitext(f)[0]
                            clean_stem = re.sub(r'_[URur]$', '', stem)
                            if is_rem or stem.endswith('_R') or stem.endswith('_U') or stem.endswith('_r') or stem.endswith('_u'):
                                rem_photos_list.append({
                                    'file': f,
                                    'stem': stem,
                                    'clean_stem': clean_stem.upper(),
                                    'ext': ext,
                                    'root': root,
                                    'date_folder': d_name,
                                    'full_path': os.path.join(root, f)
                                })
                            else:
                                orig_photos_by_date.setdefault(d_name, set()).add(clean_stem.upper())

                all_orig_global = set()
                for s in orig_photos_by_date.values():
                    all_orig_global.update(s)

                matched_u = 0
                not_matched_r = 0
                all_items = []
                date_summaries = {}

                for itm in rem_photos_list:
                    d_name = itm['date_folder']
                    orig_set = orig_photos_by_date.get(d_name, all_orig_global)
                    is_used = itm['clean_stem'] in orig_set
                    status = 'Used' if is_used else 'Missing'
                    suffix = 'U' if is_used else 'R'
                    base_clean = re.sub(r'_[URur]$', '', itm['stem'])
                    target_name = f"{base_clean}_{suffix}{itm['ext']}"

                    if is_used:
                        matched_u += 1
                    else:
                        not_matched_r += 1

                    date_summaries[d_name] = date_summaries.get(d_name, 0) + 1

                    all_items.append({
                        'id': itm['full_path'],
                        'date_folder_name': d_name,
                        'current_filename': itm['file'],
                        'proposed_new_filename': target_name,
                        'original_name': itm['file'],
                        'target_name': target_name,
                        'full_path': itm['full_path'],
                        'target_path': os.path.join(itm['root'], target_name),
                        'folder': d_name,
                        'status': status,
                        'status_text': 'Matched (_U)' if is_used else 'Not Matched (_R)'
                    })

                self._last_scan_result = {
                    "success": True,
                    "directory": clean_path,
                    "dateFoldersCount": len(date_summaries),
                    "dateFoldersList": list(date_summaries.keys()),
                    "date_summaries": date_summaries,
                    "totalRemaining": len(all_items),
                    "total_remaining": len(all_items),
                    "totalUsed": matched_u,
                    "total_used": matched_u,
                    "totalMissing": not_matched_r,
                    "total_missing": not_matched_r,
                    "totalDuplicates": 0,
                    "total_duplicates": 0,
                    "totalErrors": 0,
                    "total_errors": 0,
                    "items": all_items,
                    "all_items": all_items
                }
                self.scanCompleted.emit(self._last_scan_result)
            except Exception as e:
                err_str = f"Scan failed: {str(e)}"
                self.error.emit(err_str)
                self.scanError.emit(err_str)
            finally:
                self._scanning = False
                self.scanningChanged.emit()

        threading.Thread(target=worker, daemon=True).start()

    @Slot()
    def rename(self):
        self.executeRename(None)

    @Slot("QVariant")
    def executeRename(self, payload=None):
        payload = to_py_variant(payload)
        items = []
        if payload and isinstance(payload, dict):
            items = payload.get("items", [])
        elif payload and isinstance(payload, list):
            items = payload
        elif self._last_scan_result:
            items = self._last_scan_result.get("all_items") or self._last_scan_result.get("items") or []

        if not items:
            self.error.emit("No remaining photo items found to rename.")
            return

        self._renaming = True
        self.renamingChanged.emit()

        def worker():
            succ = 0
            fail = 0
            errors = []
            try:
                for itm in items:
                    src = itm.get("full_path", "")
                    tgt_name = itm.get("proposed_new_filename") or itm.get("target_name", "")
                    if src and tgt_name and os.path.exists(src):
                        dest = os.path.join(os.path.dirname(src), tgt_name)
                        if os.path.abspath(src) != os.path.abspath(dest):
                            try:
                                if os.path.exists(dest):
                                    os.remove(dest)
                                os.rename(src, dest)
                                itm["full_path"] = dest
                                itm["current_filename"] = tgt_name
                                succ += 1
                            except Exception as ex:
                                errors.append(f"{os.path.basename(src)}: {ex}")
                                fail += 1
                        else:
                            succ += 1
                    else:
                        fail += 1

                # Auto-export report into Document folder
                try:
                    target_paths = [t.get("full_path") for t in items if t.get("full_path")]
                    if target_paths:
                        base_dir = os.path.commonpath([os.path.dirname(p) for p in target_paths])
                        doc_dir = os.path.join(base_dir, "Document")
                        os.makedirs(doc_dir, exist_ok=True)
                        ts = time.strftime("%Y%m%d_%H%M%S")
                        
                        # 1. Export Excel (.xlsx)
                        xlsx_p = os.path.join(doc_dir, f"Photo_Match_Report_{ts}.xlsx")
                        try:
                            import openpyxl
                            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                            wb = openpyxl.Workbook()
                            ws = wb.active
                            ws.title = "Photo Match Report"

                            ws.merge_cells("A1:F1")
                            ws["A1"] = "SEQUORA Studio — Photo Match Report"
                            ws["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
                            ws["A1"].fill = PatternFill(start_color="4B2C82", end_color="4B2C82", fill_type="solid")
                            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                            ws.row_dimensions[1].height = 28

                            hdr_fill = PatternFill(start_color="7C5CBF", end_color="7C5CBF", fill_type="solid")
                            hdr_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
                            thin_side = Side(style="thin", color="D9D9D9")
                            table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

                            headers = ["SR NO", "DATE FOLDER", "ORIGINAL NAME", "NEW NAME", "STATUS", "FULL PATH"]
                            ws.row_dimensions[3].height = 22
                            for col_idx, hdr in enumerate(headers, 1):
                                cell = ws.cell(row=3, column=col_idx, value=hdr)
                                cell.font = hdr_font
                                cell.fill = hdr_fill
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                cell.border = table_border

                            zebra_fill = PatternFill(start_color="F7F5FC", end_color="F7F5FC", fill_type="solid")
                            for row_idx, itm in enumerate(items, 4):
                                is_even = (row_idx % 2 == 0)
                                row_data = [
                                    row_idx - 3,
                                    itm.get("date_folder_name", ""),
                                    itm.get("original_name", ""),
                                    itm.get("proposed_new_filename", ""),
                                    itm.get("status", ""),
                                    itm.get("full_path", "")
                                ]
                                ws.row_dimensions[row_idx].height = 20
                                for col_idx, val in enumerate(row_data, 1):
                                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                                    cell.font = Font(name="Segoe UI", size=9, color="222222")
                                    if is_even:
                                        cell.fill = zebra_fill
                                    cell.border = table_border
                                    if col_idx in (1, 5):
                                        cell.alignment = Alignment(horizontal="center", vertical="center")
                                    else:
                                        cell.alignment = Alignment(horizontal="left", vertical="center")

                            for col in ws.columns:
                                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                                max_len = max(len(str(c.value or "")) for c in col if c.row > 1) if col else 12
                                ws.column_dimensions[col_letter].width = max(12, min(max_len + 3, 60))

                            wb.save(xlsx_p)
                        except Exception as ex_wb:
                            print(f"Photo report Excel export notice: {ex_wb}")

                        # 2. Export CSV Fallback
                        csv_p = os.path.join(doc_dir, f"Photo_Match_Report_{ts}.csv")
                        with open(csv_p, "w", newline="", encoding="utf-8-sig") as f_csv:
                            writer = csv.writer(f_csv)
                            writer.writerow(["SR NO", "DATE FOLDER", "ORIGINAL NAME", "NEW NAME", "STATUS", "FULL PATH"])
                            for idx, itm in enumerate(items, 1):
                                writer.writerow([
                                    idx,
                                    itm.get("date_folder_name", ""),
                                    itm.get("original_name", ""),
                                    itm.get("proposed_new_filename", ""),
                                    itm.get("status", ""),
                                    itm.get("full_path", "")
                                ])
                        self.reportExported.emit("excel", xlsx_p if os.path.exists(xlsx_p) else csv_p)
                except Exception as exp:
                    print(f"Photo report export notice: {exp}")

                updated_report = dict(self._last_scan_result) if self._last_scan_result else {}
                updated_report["all_items"] = items
                updated_report["items"] = items
                if 'xlsx_p' in locals() and os.path.exists(xlsx_p):
                    updated_report["autoExportedReportPath"] = xlsx_p
                elif 'csv_p' in locals():
                    updated_report["autoExportedReportPath"] = csv_p
                if 'doc_dir' in locals():
                    updated_report["documentFolderPath"] = doc_dir
                self._last_scan_result = updated_report

                self.renameCompleted.emit(succ, fail, errors, updated_report)
            except Exception as e:
                self.error.emit(f"Rename failed: {str(e)}")
            finally:
                self._renaming = False
                self.renamingChanged.emit()

        threading.Thread(target=worker, daemon=True).start()

    @Slot(str, str)
    def exportReport(self, format_type, export_path):
        try:
            clean_path = export_path.replace("file:///", "").strip()
            items = self._last_scan_result.get("all_items", []) if self._last_scan_result else []
            os.makedirs(os.path.dirname(clean_path), exist_ok=True)
            with open(clean_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["SR NO", "DATE FOLDER", "ORIGINAL FILENAME", "PROPOSED NEW NAME", "STATUS", "FILE PATH"])
                for idx, itm in enumerate(items, 1):
                    writer.writerow([
                        idx,
                        itm.get("date_folder_name", ""),
                        itm.get("current_filename", ""),
                        itm.get("proposed_new_filename", ""),
                        itm.get("status_text", itm.get("status", "")),
                        itm.get("full_path", "")
                    ])
            self.reportExported.emit(format_type, clean_path)
        except Exception as e:
            self.error.emit(f"Failed to export report: {str(e)}")


class VideoTransferEngineWrapper(QObject):
    scanningChanged = Signal()
    transferringChanged = Signal()
    undoingChanged = Signal()
    matchedItemsChanged = Signal(list)
    unmatchedItemsChanged = Signal(list)
    dateFoldersChanged = Signal(list)
    photoFoldersByDateChanged = Signal(dict)
    scanResultReady = Signal(bool, str)
    dateMismatch = Signal(list, list)
    dryRunResult = Signal(dict)
    transferCompleted = Signal(dict, list)
    transferError = Signal(str)
    undoCompleted = Signal(int, int)
    error = Signal(str)
    collisionPolicyChanged = Signal()
    transferModeChanged = Signal(str)
    transferProgress = Signal()
    checkingDryRunChanged = Signal()

    VIDEO_EXTS = {
        '.mp4', '.mov', '.avi', '.mkv', '.mts', '.m2ts', '.mxf', '.wmv',
        '.flv', '.webm', '.m4v', '.mpg', '.mpeg', '.3gp', '.mod', '.tod'
    }

    IMAGE_EXTS = {
        '.jpg', '.jpeg', '.png', '.webp', '.thm', '.bmp', '.tiff', '.tif', '.heic', '.avif', '.gif'
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        self._scanning = False
        self._transferring = False
        self._undoing = False
        self._collisionPolicy = "keepBoth"
        self._transferMode = "copy"
        self._transferPercent = 0
        self._speedMbps = 0.0
        self._etaSec = 0
        self._currentFile = ""
        self._checkingDryRun = False

    @Property(bool, notify=scanningChanged)
    def scanning(self): return self._scanning

    @Property(bool, notify=transferringChanged)
    def transferring(self): return self._transferring

    @Property(bool, notify=undoingChanged)
    def undoing(self): return self._undoing

    @Property(str, notify=collisionPolicyChanged)
    def collisionPolicy(self): return self._collisionPolicy

    @collisionPolicy.setter
    def collisionPolicy(self, val):
        self._collisionPolicy = val
        self.collisionPolicyChanged.emit()

    @Property(str, notify=transferModeChanged)
    def transferMode(self): return self._transferMode

    @transferMode.setter
    def transferMode(self, val):
        self._transferMode = val
        self.transferModeChanged.emit(val)

    @Property(int, notify=transferProgress)
    def transferPercent(self): return self._transferPercent

    @Property(float, notify=transferProgress)
    def speedMbps(self): return self._speedMbps

    @Property(int, notify=transferProgress)
    def etaSec(self): return self._etaSec

    @Property(str, notify=transferProgress)
    def currentFile(self): return self._currentFile

    @Property(bool, notify=checkingDryRunChanged)
    def checkingDryRun(self): return self._checkingDryRun

    @staticmethod
    def normalize_key(s: str) -> str:
        return re.sub(r'[\s_\-]+', ' ', s.lower()).strip()

    @staticmethod
    def parse_sequence_name(raw_name: str) -> dict:
        ext = os.path.splitext(raw_name)[1]
        name_without_ext = os.path.splitext(raw_name)[0].strip() if ext else raw_name.strip()
        
        # 1. Strip trailing video/thumbnail suffixes (-its-v, _v, -v, _p, -p)
        clean = re.sub(r'[-_](?:its|v|p)+(?:[-_]v|[-_]p)?$', '', name_without_ext, flags=re.I).strip()
        # 2. Strip leading date format if present (e.g. 2026-09-01_ or 1448-03-12_)
        clean = re.sub(r'^\d{4}[-_.]\d{2}[-_.]\d{2}[\s_-]*', '', clean).strip()

        # 3. Check for camera clip prefix like '001 01-Q', '002 16-Q', 'C0001_01-Q', 'CAM1_01-Q', '01_01-Q'
        m_cam = re.match(r'^(?:[a-zA-Z]{0,4}\d{2,5})[\s_\-]+(\d+[a-zA-Z]?[\s_\-]+[a-zA-Z0-9]+(?:[\s_\-].*)?)$', clean)
        if m_cam:
            clean = m_cam.group(1).strip()

        # 4. Extract sequence number (e.g. '01', '16', '17', '01a') and code part (e.g. 'Q', 'MZ', 'Z', 'BZ', 'BQ', 'BM', 'M', 'KG', 'N', 'Reception')
        m_seq = re.match(r'^(\d+[a-zA-Z]?)(?:[\s_\-]+([a-zA-Z0-9]+))?', clean)
        sequence = m_seq.group(1) if m_seq else ""
        code_part = m_seq.group(2) if (m_seq and m_seq.group(2)) else ""

        # Normalize sequence numbers (e.g., '1' -> '01', '01' -> '01')
        norm_seq = f"{int(sequence):02d}" if sequence.isdigit() else sequence.lower()
        norm_code = code_part.lower()
        token = f"{norm_seq}-{norm_code}" if (norm_seq and norm_code) else (norm_seq or clean.lower())

        prefix = code_part.upper() if code_part else ""

        return {
            "raw": raw_name,
            "cleanName": clean,
            "sequence": norm_seq,
            "namePart": code_part or clean,
            "prefix": prefix,
            "code": norm_code,
            "token": token
        }

    @staticmethod
    def get_shot_code(raw_name: str) -> str:
        parsed = VideoTransferEngineWrapper.parse_sequence_name(raw_name)
        return parsed["token"]

    @staticmethod
    def is_match_pair(vid_parsed: dict, pf_parsed: dict) -> bool:
        v_seq = vid_parsed.get("sequence", "")
        p_seq = pf_parsed.get("sequence", "")
        v_code = vid_parsed.get("code", "")
        p_code = pf_parsed.get("code", "")
        v_tok = vid_parsed.get("token", "")
        p_tok = pf_parsed.get("token", "")
        v_clean = vid_parsed.get("cleanName", "").lower()
        p_clean = pf_parsed.get("cleanName", "").lower()

        # CRITICAL RULE: If BOTH have sequence numbers and they do NOT match, REJECT!
        if v_seq and p_seq and v_seq != p_seq:
            return False

        # 1. Exact clean name match
        if v_clean and p_clean and v_clean == p_clean:
            return True

        # 2. Token match (e.g. '01-q' == '01-q', '16-q' == '16-q', '02-mz' == '02-mz')
        if v_tok and p_tok and v_tok == p_tok:
            return True

        # 3. Matching sequence + matching code
        if v_seq and p_seq and v_seq == p_seq:
            if v_code and p_code and v_code == p_code:
                return True
            if not v_code or not p_code:
                return True

        # 4. Normalized key match when neither has conflicting sequence
        v_norm = re.sub(r'[\s_\-]+', ' ', v_clean).strip()
        p_norm = re.sub(r'[\s_\-]+', ' ', p_clean).strip()
        if v_norm and p_norm and v_norm == p_norm:
            return True

        return False

    @staticmethod
    def strip_video_suffix(raw_name: str) -> str:
        ext = os.path.splitext(raw_name)[1]
        stem = os.path.splitext(raw_name)[0].strip() if ext else raw_name.strip()
        stem = re.sub(r'[-_]its[-_][vp]$', '', stem, flags=re.I)
        stem = re.sub(r'[-_][vp]$', '', stem, flags=re.I)
        return stem.strip()

    @staticmethod
    def is_ignored_event_folder(name: str) -> bool:
        """
        Returns True if folder name represents event abbreviations MQ, MF, or H
        which never have video files and should be ignored from mismatch reporting:
        - 01-MQ, 02-MQ, 17-MQ, 18-01MQ-50423626
        - 01-MF, 02-MF, 17-MF, 18-01MF-50423626
        - 01-H, 10-H, 11-H, 17-H, 18-01H-50423626
        """
        if not name:
            return False
        clean = name.strip()
        return bool(re.search(r'(?:^|[-_\s])\d*(?:MQ|MF|H)(?:[-_\s]|$)', clean, re.IGNORECASE))

    @staticmethod
    def extract_item_code(raw_name: str) -> str:
        parsed = VideoTransferEngineWrapper.parse_sequence_name(raw_name)
        return parsed["token"]

    @classmethod
    def find_video_thumbnail(cls, video_path: str):
        if not video_path or not os.path.isfile(video_path):
            return None
        v_dir = os.path.dirname(video_path)
        vid_stem = os.path.splitext(os.path.basename(video_path))[0].lower().strip()
        base_no_suffix = re.sub(r'[-_][vp]$', '', vid_stem).strip()
        vid_shot = cls.get_shot_code(os.path.basename(video_path))
        try:
            for fname in os.listdir(v_dir):
                fp = os.path.join(v_dir, fname)
                if not os.path.isfile(fp):
                    continue
                ext = os.path.splitext(fname)[1].lower()
                if ext not in cls.IMAGE_EXTS:
                    continue
                name_lower = fname.lower()
                ent_stem = os.path.splitext(fname)[0].lower().strip()
                ent_shot = cls.get_shot_code(fname)
                if name_lower.endswith('_v.jpg') or name_lower.endswith('_v.png') or name_lower.endswith('_v.jpeg'):
                    if ent_shot == vid_shot or ent_stem.startswith(base_no_suffix):
                        return {"path": fp, "name": fname}
                if ent_shot == vid_shot or ent_stem == vid_stem:
                    return {"path": fp, "name": fname}
        except Exception:
            pass
        return None

    @classmethod
    def find_photo_thumbnail(cls, photo_folder_path: str, video_path: str = None):
        if photo_folder_path and os.path.isdir(photo_folder_path):
            try:
                for fname in os.listdir(photo_folder_path):
                    fp = os.path.join(photo_folder_path, fname)
                    if not os.path.isfile(fp):
                        continue
                    ext = os.path.splitext(fname)[1].lower()
                    if ext not in cls.IMAGE_EXTS:
                        continue
                    name_lower = fname.lower()
                    if name_lower.endswith('_p.jpg') or name_lower.endswith('_p.png') or name_lower.endswith('_p.jpeg'):
                        return {"path": fp, "name": fname}
            except Exception:
                pass

        if video_path and os.path.isfile(video_path):
            v_dir = os.path.dirname(video_path)
            v_shot = cls.get_shot_code(video_path)
            v_stem_lower = os.path.splitext(os.path.basename(video_path))[0].lower().strip()
            v_base_no_suffix = re.sub(r'[-_][vp]$', '', v_stem_lower).strip()
            try:
                for fname in os.listdir(v_dir):
                    fp = os.path.join(v_dir, fname)
                    if not os.path.isfile(fp):
                        continue
                    ext = os.path.splitext(fname)[1].lower()
                    if ext not in cls.IMAGE_EXTS:
                        continue
                    name_lower = fname.lower()
                    ent_shot = cls.get_shot_code(fname)
                    if (name_lower.endswith('_p.jpg') or name_lower.endswith('_p.png') or name_lower.endswith('_p.jpeg')) and \
                       (ent_shot == v_shot or name_lower.startswith(v_base_no_suffix)):
                        return {"path": fp, "name": fname}
            except Exception:
                pass
        return None

    @Slot("QVariant")
    def scan(self, params):
        self.scanFolder(params)

    @Slot(str, str)
    def scan(self, video_dir, photo_dir):
        self.scanFolder({"videoFolderPath": video_dir, "photoFolderPath": photo_dir})

    @Slot("QVariant")
    def scanFolder(self, params):
        params = to_py_variant(params)
        v_path = ""
        p_path = ""

        if isinstance(params, str):
            v_path = params
            p_path = params
        elif isinstance(params, dict):
            master = params.get("masterFolderPath", "")
            if master:
                master_clean = os.path.abspath(master.replace("file:///", "").strip())
                if os.path.exists(master_clean):
                    try:
                        subdirs = [d.name for d in os.scandir(master_clean) if d.is_dir()]
                        v_sub = next((d for d in subdirs if re.search(r'video|dvd|vid', d, re.I)), None)
                        p_sub = next((d for d in subdirs if re.search(r'photo|pic', d, re.I)), None)
                        v_path = os.path.join(master_clean, v_sub) if v_sub else master_clean
                        p_path = os.path.join(master_clean, p_sub) if p_sub else master_clean
                    except Exception:
                        v_path = master_clean
                        p_path = master_clean
                else:
                    v_path = master_clean
                    p_path = master_clean
            else:
                v_path = params.get("videoFolderPath", "")
                p_path = params.get("photoFolderPath", "")

        if not v_path or not p_path:
            return

        v_clean = os.path.abspath(v_path.replace("file:///", "").strip())
        p_clean = os.path.abspath(p_path.replace("file:///", "").strip())

        if not os.path.exists(v_clean):
            err = f"Video directory does not exist: {v_clean}"
            self.error.emit(err)
            self.scanResultReady.emit(False, err)
            return

        if not os.path.exists(p_clean):
            err = f"Photo directory does not exist: {p_clean}"
            self.error.emit(err)
            self.scanResultReady.emit(False, err)
            return

        self._scanning = True
        self.scanningChanged.emit()

        def scan_worker():
            all_discovered_videos = []
            all_photo_subfolders = []
            try:
                def is_date_folder_name(name):
                    if re.search(r'\d{4}[-_.]\d{2}[-_.]\d{2}', name) or re.search(r'\d{2}[-_.]\d{2}[-_.]\d{4}', name):
                        return True
                    if re.match(r'^(?:day|date|event|session)[\s_\-]*\d+', name, re.I):
                        return True
                    return False
    
                def is_internal_subfolder(name):
                    return bool(re.match(r'^(?:raw|jpg|jpeg|png|cr2|cr3|nef|arw|dng|edited|selected|selection|hires|lowres|exports?|thumbs?|thumbnails?)$', name, re.I))
    
                # 1. Discover ALL Video files and their date folders under v_clean
                def scan_video_dir(current_dir, depth):
                    if depth > 5:
                        return
                    try:
                        for entry in os.scandir(current_dir):
                            if entry.is_dir():
                                scan_video_dir(entry.path, depth + 1)
                            elif entry.is_file():
                                ext = os.path.splitext(entry.name)[1].lower()
                                if ext in self.VIDEO_EXTS:
                                    rel = os.path.relpath(entry.path, v_clean)
                                    parts = [p for p in rel.replace('\\', '/').split('/') if p]
                                    date_folder_name = os.path.basename(v_clean)
                                    
                                    for p in parts[:-1]:
                                        if is_date_folder_name(p):
                                            date_folder_name = p
                                            break
                                    else:
                                        if len(parts) > 1:
                                            top_folder = parts[0]
                                            if re.match(r'^(?:dvd|videos?)$', top_folder, re.I):
                                                date_folder_name = parts[1] if len(parts) > 2 else os.path.basename(v_clean)
                                            else:
                                                date_folder_name = top_folder
                                    
                                    thumb = self.find_video_thumbnail(entry.path)
                                    all_discovered_videos.append({
                                        "filePath": entry.path,
                                        "fileName": entry.name,
                                        "dateFolderName": date_folder_name,
                                        "parsed": self.parse_sequence_name(entry.name),
                                        "shotCode": self.get_shot_code(entry.name),
                                        "thumbnailPath": thumb["path"] if thumb else "",
                                        "thumbnailName": thumb["name"] if thumb else ""
                                    })
                    except Exception:
                        pass
    
                scan_video_dir(v_clean, 0)
                QApplication.processEvents()
    
                # 2. Discover ALL Photo subfolders and their date folders under p_clean
                def scan_photo_dir(current_dir, depth):
                    if depth > 6:
                        return
                    try:
                        for entry in os.scandir(current_dir):
                            if entry.is_dir():
                                if is_internal_subfolder(entry.name):
                                    continue
    
                                full_path = entry.path
                                rel = os.path.relpath(full_path, p_clean)
                                parts = [p for p in rel.replace('\\', '/').split('/') if p]
    
                                date_folder_name = os.path.basename(p_clean)
                                for p in parts:
                                    if is_date_folder_name(p):
                                        date_folder_name = p
                                        break
                                else:
                                    if len(parts) > 1:
                                        date_folder_name = parts[0]
    
                                if not (len(parts) == 1 and is_date_folder_name(parts[0])):
                                    all_photo_subfolders.append({
                                        "path": full_path,
                                        "name": entry.name,
                                        "dateFolderName": date_folder_name,
                                        "parsed": self.parse_sequence_name(entry.name),
                                        "shotCode": self.get_shot_code(entry.name),
                                        "paired": False
                                    })
                                scan_photo_dir(full_path, depth + 1)
                    except Exception:
                        pass
    
                scan_photo_dir(p_clean, 0)
                QApplication.processEvents()
    
                def is_generic_folder(name):
                    return bool(re.match(r'^(?:video|videos|dvd|vid|photo|photos|pic|pics|media|raw|edited|selection)$', name, re.I))
    
                # 3. Extract unique date folder names for Video and Photo
                video_date_names = list(dict.fromkeys(v["dateFolderName"] for v in all_discovered_videos))
                photo_date_names = list(dict.fromkeys(p["dateFolderName"] for p in all_photo_subfolders))
    
                v_date_list = video_date_names if video_date_names else [os.path.basename(v_clean)]
                p_date_list = photo_date_names if photo_date_names else [os.path.basename(p_clean)]
    
                matching_date_keys = [
                    v_date for v_date in v_date_list
                    if any(self.normalize_key(v_date) == self.normalize_key(p_date) for p_date in p_date_list)
                ]
    
                is_single_top = (len(v_date_list) <= 1 and len(p_date_list) <= 1)
                v_first = v_date_list[0] if v_date_list else ""
                p_first = p_date_list[0] if p_date_list else ""
                is_compatible = (len(matching_date_keys) > 0) or (
                    is_single_top and (is_generic_folder(v_first) or is_generic_folder(p_first) or not (re.search(r'\d{4}', v_first) and re.search(r'\d{4}', p_first)))
                )
    
                if not is_compatible:
                    self.dateMismatch.emit(v_date_list, p_date_list)
                    err_msg = f"Selected folders are of 2 different names!\nVideo Date Folder: {', '.join(v_date_list)}\nPhoto Date Folder: {', '.join(p_date_list)}"
                    self.scanResultReady.emit(False, err_msg)
                    return
    
                # Populate photoFoldersByDate lookup
                photo_folders_by_date = {}
                all_photo_list = []
                for sf in all_photo_subfolders:
                    d_name = sf["dateFolderName"]
                    if d_name not in photo_folders_by_date:
                        photo_folders_by_date[d_name] = []
                    photo_folders_by_date[d_name].append({"name": sf["name"], "path": sf["path"]})
                    all_photo_list.append({"name": f"{d_name} / {sf['name']}", "path": sf["path"]})
                photo_folders_by_date["All Folders"] = all_photo_list
    
                # 4. Pair Video Files <-> Photo Subfolders ONLY WITHIN MATCHING DATE FOLDERS
                matched = []
                unmatched = []
    
                for vid in all_discovered_videos:
                    pf_match = None
                    for pf in all_photo_subfolders:
                        if pf["paired"]:
                            continue
                        if len(video_date_names) > 1 or len(photo_date_names) > 1:
                            if self.normalize_key(vid["dateFolderName"]) != self.normalize_key(pf["dateFolderName"]):
                                continue
    
                        if self.is_match_pair(vid["parsed"], pf["parsed"]):
                            pf_match = pf
                            break
    
                    if pf_match:
                        v_thumb = self.find_video_thumbnail(vid["filePath"])
                        p_thumb = self.find_photo_thumbnail(pf_match["path"], vid["filePath"])
    
                        matched.append({
                            "id": f"{vid['filePath']}_{pf_match['path']}",
                            "videoPath": vid["filePath"],
                            "videoName": vid["fileName"],
                            "thumbnailPath": v_thumb["path"] if v_thumb else vid.get("thumbnailPath", ""),
                            "thumbnailName": v_thumb["name"] if v_thumb else vid.get("thumbnailName", ""),
                            "videoThumbnailPath": v_thumb["path"] if v_thumb else vid.get("thumbnailPath", ""),
                            "videoThumbnailName": v_thumb["name"] if v_thumb else vid.get("thumbnailName", ""),
                            "photoThumbnailPath": p_thumb["path"] if p_thumb else "",
                            "photoThumbnailName": p_thumb["name"] if p_thumb else "",
                            "photoFolderPath": pf_match["path"],
                            "photoFolderName": pf_match["name"],
                            "dateFolderName": pf_match["dateFolderName"] or vid["dateFolderName"],
                            "sequence": vid["parsed"]["sequence"] or pf_match["parsed"]["sequence"],
                            "prefix": vid["parsed"]["prefix"] or pf_match["parsed"]["prefix"],
                            "namePart": vid["parsed"]["namePart"],
                            "status": "APPROVED",
                            "isMatched": True
                        })
                        pf_match["paired"] = True
                    else:
                        v_thumb = self.find_video_thumbnail(vid["filePath"])
                        unmatched.append({
                            "id": f"unmatched_vid_{vid['filePath']}",
                            "videoPath": vid["filePath"],
                            "videoName": vid["fileName"],
                            "thumbnailPath": v_thumb["path"] if v_thumb else vid.get("thumbnailPath", ""),
                            "thumbnailName": v_thumb["name"] if v_thumb else vid.get("thumbnailName", ""),
                            "videoThumbnailPath": v_thumb["path"] if v_thumb else vid.get("thumbnailPath", ""),
                            "videoThumbnailName": v_thumb["name"] if v_thumb else vid.get("thumbnailName", ""),
                            "dateFolderName": vid["dateFolderName"],
                            "sequence": vid["parsed"]["sequence"],
                            "prefix": vid["parsed"]["prefix"],
                            "namePart": vid["parsed"]["namePart"],
                            "reason": f'No matching photo subfolder found for video filename "{vid["fileName"]}" in date folder "{vid["dateFolderName"]}"',
                            "status": "MISMATCH",
                            "isMatched": False
                        })
    
                matched_paths = {m["photoFolderPath"] for m in matched}
    
                for pf in all_photo_subfolders:
                    if not pf["paired"]:
                        p_path = pf["path"]
                        p_name = pf["name"]

                        # Skip photo folders with MQ, MF, or H event abbreviations as they have no video files
                        if self.is_ignored_event_folder(p_name) or self.is_ignored_event_folder(os.path.basename(p_path)):
                            continue

                        # Skip if inside already matched parent or is container of a matched folder
                        if any(p_path.startswith(mp + os.sep) for mp in matched_paths):
                            continue
                        if any(mp.startswith(p_path + os.sep) for mp in matched_paths):
                            continue
                        # Must look like a sequence folder (e.g. 01-Z, 12-KG, 01a-N)
                        if not re.match(r'^\d+[a-zA-Z]*[\s_\-]', pf["name"].strip()):
                            continue
    
                        unmatched.append({
                            "id": f"unmatched_pf_{pf['path']}",
                            "photoFolderPath": pf["path"],
                            "photoFolderName": pf["name"],
                            "dateFolderName": pf["dateFolderName"],
                            "sequence": pf["parsed"]["sequence"],
                            "prefix": pf["parsed"]["prefix"],
                            "namePart": pf["parsed"]["namePart"],
                            "reason": f'Photo subfolder "{pf["name"]}" in date folder "{pf["dateFolderName"]}" has no matching video file',
                            "status": "MISMATCH",
                            "isMatched": False
                        })
    
                # 5. Build dateFolders HUD summary
                date_folder_map = {}
                for m in matched:
                    d_name = m.get("dateFolderName") or "General"
                    curr = date_folder_map.setdefault(d_name, {"videoCount": 0, "photoFolderCount": 0})
                    curr["videoCount"] += 1
    
                for pf in all_photo_subfolders:
                    d_name = pf["dateFolderName"]
                    curr = date_folder_map.setdefault(d_name, {"videoCount": 0, "photoFolderCount": 0})
                    curr["photoFolderCount"] += 1
    
                date_folders = [
                    {
                        "name": k,
                        "videoCount": v["videoCount"],
                        "photoFolderCount": v["photoFolderCount"]
                    }
                    for k, v in date_folder_map.items()
                ]
    
                photo_folders_by_date = {}
                all_pf_list = []
                for pf in all_photo_subfolders:
                    if not re.match(r'^\d+[a-zA-Z]*[\s_\-]', pf["name"].strip()):
                        continue
                    d_name = pf["dateFolderName"]
                    item = {
                        "path": pf["path"],
                        "name": pf["name"],
                        "dateFolderName": d_name
                    }
                    photo_folders_by_date.setdefault(d_name, []).append(item)
                    all_pf_list.append(item)
                photo_folders_by_date["All Folders"] = all_pf_list
                self._matched = matched
                self._unmatched = unmatched
                self.matchedItemsChanged.emit(matched)
                self.unmatchedItemsChanged.emit(unmatched)
                self.dateFoldersChanged.emit(date_folders)
                self.photoFoldersByDateChanged.emit(photo_folders_by_date)
                self.scanResultReady.emit(True, "")
    
            except Exception as e:
                self.error.emit(f"Scan failed: {str(e)}")
                self.scanResultReady.emit(False, str(e))
            finally:
                self._scanning = False
                self.scanningChanged.emit()

        threading.Thread(target=scan_worker, daemon=True).start()

    @Slot(str, str)
    def linkVideoToPhotoFolder(self, video_path, photo_folder_path):
        """
        Manually links an unmatched video clip (or re-links an existing clip)
        to a specified photo subfolder.
        """
        clean_v = os.path.abspath(video_path.replace("file:///", "").strip())
        clean_p = os.path.abspath(photo_folder_path.replace("file:///", "").strip())

        vid_item = None
        new_unmatched = []
        for u in self._unmatched:
            u_path = os.path.abspath(u.get("videoPath", "").replace("file:///", "").strip())
            if u_path == clean_v:
                vid_item = u
            else:
                new_unmatched.append(u)

        if not vid_item:
            for m in self._matched:
                m_path = os.path.abspath(m.get("videoPath", "").replace("file:///", "").strip())
                if m_path == clean_v:
                    vid_item = m
                    break

        if not vid_item:
            v_name = os.path.basename(clean_v)
            vid_item = {
                "videoPath": clean_v,
                "videoName": v_name,
                "dateFolderName": os.path.basename(os.path.dirname(clean_v)),
                "parsed": self.parse_sequence_name(v_name),
                "prefix": "",
                "namePart": v_name
            }

        p_name = os.path.basename(clean_p)
        v_thumb = self.find_video_thumbnail(clean_v)
        p_thumb = self.find_photo_thumbnail(clean_p, clean_v)

        # Remove previous match for this video
        new_matched = [m for m in self._matched if os.path.abspath(m.get("videoPath", "").replace("file:///", "").strip()) != clean_v]

        new_matched.append({
            "id": f"{clean_v}_{clean_p}",
            "videoPath": clean_v,
            "videoName": vid_item.get("videoName", os.path.basename(clean_v)),
            "thumbnailPath": v_thumb["path"] if v_thumb else vid_item.get("thumbnailPath", ""),
            "thumbnailName": v_thumb["name"] if v_thumb else vid_item.get("thumbnailName", ""),
            "videoThumbnailPath": v_thumb["path"] if v_thumb else vid_item.get("videoThumbnailPath", ""),
            "videoThumbnailName": v_thumb["name"] if v_thumb else vid_item.get("videoThumbnailName", ""),
            "photoThumbnailPath": p_thumb["path"] if p_thumb else "",
            "photoThumbnailName": p_thumb["name"] if p_thumb else "",
            "photoFolderPath": clean_p,
            "photoFolderName": p_name,
            "dateFolderName": vid_item.get("dateFolderName") or os.path.basename(os.path.dirname(clean_p)),
            "sequence": vid_item.get("sequence", ""),
            "prefix": vid_item.get("prefix", ""),
            "namePart": vid_item.get("namePart", ""),
            "status": "APPROVED",
            "isMatched": True
        })

        self._unmatched = new_unmatched
        self._matched = new_matched
        self.matchedItemsChanged.emit(self._matched)
        self.unmatchedItemsChanged.emit(self._unmatched)

    @Slot(str, str, str)
    def scanDirectories(self, video_dir, photo_dir, master_dir=""):
        payload = {}
        if master_dir:
            payload["masterFolderPath"] = master_dir
        else:
            payload["videoFolderPath"] = video_dir
            payload["photoFolderPath"] = photo_dir
        self.scanFolder(payload)

    @Slot("QVariant")
    def checkDryRun(self, items):
        items = to_py_variant(items)
        if not items:
            self.dryRunResult.emit({"success": False, "error": "No items to check."})
            return

        self._checkingDryRun = True
        self.checkingDryRunChanged.emit()
        try:
            total_bytes = 0
            file_count = 0
            check_dir = None

            for itm in items:
                v_path = itm.get("videoPath", "")
                t_path = itm.get("thumbnailPath", "")
                p_folder = itm.get("photoFolderPath", "")

                if v_path and os.path.isfile(v_path):
                    total_bytes += os.path.getsize(v_path)
                    file_count += 1
                if t_path and os.path.isfile(t_path):
                    total_bytes += os.path.getsize(t_path)
                    file_count += 1

                if not check_dir and p_folder:
                    check_dir = p_folder if os.path.isdir(p_folder) else os.path.dirname(p_folder)

            free_bytes = None
            if check_dir and os.path.exists(check_dir):
                try:
                    usage = shutil.disk_usage(check_dir)
                    free_bytes = usage.free
                except Exception:
                    pass

            res = {
                "success": True,
                "totalBytes": total_bytes,
                "fileCount": file_count,
                "freeBytes": free_bytes,
                "hasSufficientSpace": True if free_bytes is None else (free_bytes >= total_bytes)
            }
            self.dryRunResult.emit(res)
        except Exception as e:
            self.dryRunResult.emit({"success": False, "error": str(e)})
        finally:
            self._checkingDryRun = False
            self.checkingDryRunChanged.emit()

    @Slot("QVariant")
    def executeTransfer(self, payload):
        payload = to_py_variant(payload)
        items_to_transfer = []
        collision_policy = self._collisionPolicy
        transfer_mode = self._transferMode

        if isinstance(payload, list):
            items_to_transfer = payload
        elif isinstance(payload, dict):
            items_to_transfer = payload.get("items") or payload.get("itemsToTransfer") or payload.get("matchedClips") or []
            collision_policy = payload.get("collisionPolicy") or self._collisionPolicy
            transfer_mode = payload.get("transferMode") or self._transferMode

        if not items_to_transfer:
            self.transferError.emit("No items selected for transfer.")
            return

        self._transferring = True
        self._transferPercent = 0
        self._speedMbps = 0.0
        self._etaSec = 0
        self._currentFile = ""
        self.transferringChanged.emit()
        self.transferProgress.emit()

        def worker():
            total = len(items_to_transfer)
            done = 0
            transferred_files = []
            errors = []
            start_time = time.time()
            total_bytes_copied = 0

            total_bytes_to_copy = 0
            for item in items_to_transfer:
                vp = item.get("videoPath", "")
                tp = item.get("thumbnailPath", "") or item.get("videoThumbnailPath", "")
                if vp and os.path.isfile(vp):
                    total_bytes_to_copy += os.path.getsize(vp)
                if tp and os.path.isfile(tp):
                    total_bytes_to_copy += os.path.getsize(tp)

            def copy_chunked(src, dst):
                nonlocal total_bytes_copied
                with open(src, 'rb') as fsrc:
                    with open(dst, 'wb') as fdst:
                        while True:
                            buf = fsrc.read(4 * 1024 * 1024)
                            if not buf:
                                break
                            fdst.write(buf)
                            total_bytes_copied += len(buf)
                            elapsed = time.time() - start_time
                            if elapsed > 0.1:
                                speed_bps = total_bytes_copied / elapsed
                                self._speedMbps = round(speed_bps / (1024 * 1024), 1)
                                remaining_b = max(0, total_bytes_to_copy - total_bytes_copied)
                                self._etaSec = int(remaining_b / speed_bps) if speed_bps > 0 else 0
                                self._transferPercent = min(99, int((total_bytes_copied / max(total_bytes_to_copy, 1)) * 100))
                                self.transferProgress.emit()
                try:
                    shutil.copystat(src, dst)
                except Exception:
                    pass

            try:
                for item in items_to_transfer:
                    v_path = item.get("videoPath", "")
                    v_name = item.get("videoName") or os.path.basename(v_path)
                    p_folder = item.get("photoFolderPath", "")
                    p_name = item.get("photoFolderName") or os.path.basename(p_folder)

                    if not v_path or not p_folder:
                        continue

                    if not os.path.exists(v_path):
                        errors.append(f"Source video missing: {v_name}")
                        continue

                    os.makedirs(p_folder, exist_ok=True)
                    target_file_path = os.path.join(p_folder, os.path.basename(v_path))

                    if os.path.exists(target_file_path):
                        if collision_policy == "skip":
                            done += 1
                            continue
                        elif collision_policy == "keepBoth":
                            ext = os.path.splitext(v_path)[1]
                            name_base = os.path.splitext(os.path.basename(v_path))[0]
                            copy_idx = 1
                            while os.path.exists(target_file_path):
                                target_file_path = os.path.join(p_folder, f"{name_base}_{copy_idx}{ext}")
                                copy_idx += 1

                    self._currentFile = v_name
                    file_size = os.path.getsize(v_path)

                    # 1. Transfer Video file (Move or Copy)
                    if transfer_mode == "move":
                        same_drive = False
                        try:
                            same_drive = os.path.splitdrive(os.path.abspath(v_path))[0].lower() == os.path.splitdrive(os.path.abspath(target_file_path))[0].lower()
                        except Exception:
                            pass

                        if same_drive:
                            shutil.move(v_path, target_file_path)
                            total_bytes_copied += file_size
                        else:
                            copy_chunked(v_path, target_file_path)
                            try:
                                os.remove(v_path)
                            except Exception:
                                pass
                    else:
                        copy_chunked(v_path, target_file_path)

                    target_video_thumb_path = None
                    target_photo_thumb_path = None
                    thumb_dest_dir = os.path.dirname(p_folder)

                    # 2. Copy matching Video Thumbnail (_V.jpg)
                    v_thumb_path = item.get("videoThumbnailPath") or item.get("thumbnailPath")
                    if not v_thumb_path or not os.path.exists(v_thumb_path):
                        v_thumb_obj = self.find_video_thumbnail(v_path)
                        v_thumb_path = v_thumb_obj["path"] if v_thumb_obj else None

                    if v_thumb_path and os.path.exists(v_thumb_path):
                        orig_name = os.path.basename(v_thumb_path)
                        dest_v_thumb_path = os.path.join(thumb_dest_dir, orig_name)
                        if v_thumb_path.lower() != dest_v_thumb_path.lower() and not os.path.exists(dest_v_thumb_path):
                            os.makedirs(thumb_dest_dir, exist_ok=True)
                            shutil.copy2(v_thumb_path, dest_v_thumb_path)
                            total_bytes_copied += os.path.getsize(v_thumb_path)
                        target_video_thumb_path = dest_v_thumb_path

                    # 3. Copy matching Photo Thumbnail (_P.jpg)
                    p_thumb_path = item.get("photoThumbnailPath")
                    if not p_thumb_path or not os.path.exists(p_thumb_path):
                        p_thumb_obj = self.find_photo_thumbnail(p_folder, v_path)
                        p_thumb_path = p_thumb_obj["path"] if p_thumb_obj else None

                    if p_thumb_path and os.path.exists(p_thumb_path):
                        orig_p_name = os.path.basename(p_thumb_path)
                        dest_p_thumb_path = os.path.join(thumb_dest_dir, orig_p_name)
                        if p_thumb_path.lower() != dest_p_thumb_path.lower() and not os.path.exists(dest_p_thumb_path):
                            os.makedirs(thumb_dest_dir, exist_ok=True)
                            shutil.copy2(p_thumb_path, dest_p_thumb_path)
                            total_bytes_copied += os.path.getsize(p_thumb_path)
                        target_photo_thumb_path = dest_p_thumb_path

                    done += 1
                    transferred_files.append({
                        "videoName": v_name,
                        "photoFolderName": p_name,
                        "dateFolderName": item.get("dateFolderName", ""),
                        "sequence": item.get("sequence", ""),
                        "sourceVideoPath": v_path,
                        "targetPath": target_file_path,
                        "sourceVideoThumbPath": v_thumb_path or "",
                        "thumbnailTargetPath": target_video_thumb_path or "",
                        "sourcePhotoThumbPath": p_thumb_path or "",
                        "photoThumbnailTargetPath": target_photo_thumb_path or "",
                        "transferMode": transfer_mode
                    })

                    elapsed_sec = time.time() - start_time
                    speed_bps = total_bytes_copied / elapsed_sec if elapsed_sec > 0 else 0
                    self._speedMbps = round(speed_bps / (1024 * 1024), 1)
                    remaining_bytes = max(0, total_bytes_to_copy - total_bytes_copied)
                    self._etaSec = int(remaining_bytes / speed_bps) if speed_bps > 0 else 0
                    self._transferPercent = int((done / max(total, 1)) * 100)
                    self.transferProgress.emit()

                duration_sec = time.time() - start_time
                avg_speed_mbps = (total_bytes_copied / duration_sec) / (1024 * 1024) if duration_sec > 0 else 0

                # Automatically export styled Excel (.xlsx) undo manifest and CSV report to Document folder
                auto_csv_path = None
                auto_xlsx_path = None
                try:
                    target_paths = [t.get("targetPath") for t in transferred_files if t.get("targetPath")]
                    if target_paths:
                        dirs = [os.path.dirname(os.path.dirname(p)) for p in target_paths]
                        base_dir = os.path.commonpath(dirs) if dirs else None
                        if not base_dir or not os.path.exists(base_dir):
                            base_dir = os.path.dirname(os.path.dirname(target_paths[0])) if target_paths else '.'

                        doc_dir = os.path.join(base_dir, "Document")
                        os.makedirs(doc_dir, exist_ok=True)

                        timestamp_str = time.strftime("%Y%m%d_%H%M%S")
                        date_display = time.strftime("%Y-%m-%d %H:%M:%S")

                        # 1. Generate Excel (.xlsx) Workbook using openpyxl
                        try:
                            import openpyxl
                            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                            wb = openpyxl.Workbook()
                            ws = wb.active
                            ws.title = "Transfer & Undo Manifest"

                            # Title Header
                            ws.merge_cells("A1:M1")
                            ws["A1"] = "SEQUORA Studio — Video Transfer & Undo Manifest"
                            ws["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
                            ws["A1"].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                            ws.row_dimensions[1].height = 30

                            # Metadata Summary Rows
                            meta_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
                            meta_font = Font(name="Segoe UI", size=10, bold=True, color="333333")
                            val_font = Font(name="Segoe UI", size=10, color="222222")

                            ws["A2"] = "Export Timestamp:"
                            ws["A2"].font = meta_font
                            ws["A2"].fill = meta_fill
                            ws["B2"] = date_display
                            ws["B2"].font = val_font

                            ws["D2"] = "Total Clips:"
                            ws["D2"].font = meta_font
                            ws["D2"].fill = meta_fill
                            ws["E2"] = len(transferred_files)
                            ws["E2"].font = val_font

                            ws["G2"] = "Transfer Mode:"
                            ws["G2"].font = meta_font
                            ws["G2"].fill = meta_fill
                            ws["H2"] = transfer_mode.upper()
                            ws["H2"].font = val_font

                            ws["J2"] = "Target Document Folder:"
                            ws["J2"].font = meta_font
                            ws["J2"].fill = meta_fill
                            ws["K2"] = doc_dir
                            ws["K2"].font = val_font

                            # Table Headers
                            headers = [
                                "SR NO",
                                "DATE FOLDER",
                                "SEQUENCE",
                                "VIDEO FILENAME",
                                "SOURCE VIDEO PATH",
                                "TARGET PHOTO FOLDER",
                                "TARGET VIDEO PATH",
                                "SOURCE VIDEO THUMB",
                                "TARGET VIDEO THUMB",
                                "SOURCE PHOTO THUMB",
                                "TARGET PHOTO THUMB",
                                "TRANSFER MODE",
                                "STATUS"
                            ]

                            hdr_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
                            hdr_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
                            thin_side = Side(style="thin", color="D9D9D9")
                            table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

                            ws.row_dimensions[4].height = 24
                            for col_idx, hdr in enumerate(headers, 1):
                                cell = ws.cell(row=4, column=col_idx, value=hdr)
                                cell.font = hdr_font
                                cell.fill = hdr_fill
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                cell.border = table_border

                            # Data Rows
                            zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
                            for row_idx, tf in enumerate(transferred_files, 5):
                                row_data = [
                                    row_idx - 4,
                                    tf.get("dateFolderName", ""),
                                    tf.get("sequence", ""),
                                    tf.get("videoName", ""),
                                    tf.get("sourceVideoPath", ""),
                                    tf.get("photoFolderName", ""),
                                    tf.get("targetPath", ""),
                                    tf.get("sourceVideoThumbPath", ""),
                                    tf.get("thumbnailTargetPath", ""),
                                    tf.get("sourcePhotoThumbPath", ""),
                                    tf.get("photoThumbnailTargetPath", ""),
                                    transfer_mode.upper(),
                                    "SUCCESS"
                                ]
                                ws.row_dimensions[row_idx].height = 20
                                is_even = (row_idx % 2 == 0)
                                for col_idx, val in enumerate(row_data, 1):
                                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                                    cell.font = Font(name="Segoe UI", size=9, color="222222")
                                    if is_even:
                                        cell.fill = zebra_fill
                                    cell.border = table_border
                                    if col_idx in (1, 3, 12, 13):
                                        cell.alignment = Alignment(horizontal="center", vertical="center")
                                    else:
                                        cell.alignment = Alignment(horizontal="left", vertical="center")

                            # Auto adjust column widths
                            for col in ws.columns:
                                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                                max_len = 0
                                for cell in col:
                                    if cell.row == 1:
                                        continue
                                    val_str = str(cell.value or "")
                                    if len(val_str) > max_len:
                                        max_len = len(val_str)
                                ws.column_dimensions[col_letter].width = max(12, min(max_len + 3, 50))

                            xlsx_filename = f"Video_Transfer_Undo_Manifest_{timestamp_str}.xlsx"
                            auto_xlsx_path = os.path.join(doc_dir, xlsx_filename)
                            wb.save(auto_xlsx_path)
                        except Exception as ex_wb:
                            print(f"Excel export notice: {ex_wb}")

                        # 2. Generate CSV Fallback
                        csv_filename = f"Video_Transfer_Report_{timestamp_str}.csv"
                        auto_csv_path = os.path.join(doc_dir, csv_filename)

                        with open(auto_csv_path, "w", newline="", encoding="utf-8-sig") as f_csv:
                            writer = csv.writer(f_csv)
                            writer.writerow([
                                "SR NO",
                                "DATE FOLDER",
                                "SEQUENCE",
                                "VIDEO NAME",
                                "SOURCE VIDEO PATH",
                                "TARGET PHOTO FOLDER",
                                "TARGET VIDEO PATH",
                                "SOURCE VIDEO THUMB",
                                "TARGET VIDEO THUMB",
                                "SOURCE PHOTO THUMB",
                                "TARGET PHOTO THUMB",
                                "TRANSFER MODE",
                                "STATUS",
                                "TIMESTAMP"
                            ])
                            for idx, tf in enumerate(transferred_files, 1):
                                writer.writerow([
                                    idx,
                                    tf.get("dateFolderName", ""),
                                    tf.get("sequence", ""),
                                    tf.get("videoName", ""),
                                    tf.get("sourceVideoPath", ""),
                                    tf.get("photoFolderName", ""),
                                    tf.get("targetPath", ""),
                                    tf.get("sourceVideoThumbPath", ""),
                                    tf.get("thumbnailTargetPath", ""),
                                    tf.get("sourcePhotoThumbPath", ""),
                                    tf.get("photoThumbnailTargetPath", ""),
                                    transfer_mode.upper(),
                                    "SUCCESS",
                                    date_display
                                ])
                except Exception as exp:
                    print(f"Auto-export manifest notice: {exp}")

                report_summary = {
                    "success": True,
                    "totalClipsSelected": total,
                    "totalSuccessful": done,
                    "totalErrors": len(errors),
                    "totalBytesCopied": total_bytes_copied,
                    "durationSec": round(duration_sec, 1),
                    "avgSpeedMbps": round(avg_speed_mbps, 1),
                    "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                    "transferredFiles": transferred_files,
                    "errors": errors,
                    "autoExportedCsvPath": auto_csv_path or "",
                    "autoExportedXlsxPath": auto_xlsx_path or ""
                }

                self._transferPercent = 100
                self.transferProgress.emit()
                self.transferCompleted.emit(report_summary, transferred_files)

            except Exception as e:
                self.transferError.emit(f"Transfer execution failed: {str(e)}")
            finally:
                self._transferring = False
                self.transferringChanged.emit()

        threading.Thread(target=worker, daemon=True).start()

    @Slot("QVariant")
    def undoTransfer(self, transferred_files):
        transferred_files = to_py_variant(transferred_files)
        if not transferred_files:
            self.error.emit("No transferred items to undo.")
            return

        self._undoing = True
        self.undoingChanged.emit()

        undone_count = 0
        errors = []

        try:
            for item in transferred_files:
                target_path = item.get("targetPath", "")
                source_path = item.get("sourceVideoPath", "") or item.get("sourcePath", "")
                thumb_target_path = item.get("thumbnailTargetPath", "")
                mode = (item.get("transferMode") or "COPY").upper()

                if target_path and os.path.exists(target_path):
                    try:
                        if mode == "MOVE" and source_path:
                            # Revert by moving file BACK to source directory
                            os.makedirs(os.path.dirname(source_path), exist_ok=True)
                            shutil.move(target_path, source_path)
                            undone_count += 1
                        else:
                            os.remove(target_path)
                            undone_count += 1
                    except Exception as err:
                        errors.append(f"Failed to revert {target_path}: {str(err)}")

                if thumb_target_path and os.path.exists(thumb_target_path):
                    try:
                        os.remove(thumb_target_path)
                    except Exception as err:
                        errors.append(f"Failed to clean thumbnail {thumb_target_path}: {str(err)}")

            self.undoCompleted.emit(undone_count, len(errors))
        except Exception as e:
            self.error.emit(f"Undo failed: {str(e)}")
        finally:
            self._undoing = False
            self.undoingChanged.emit()


class PVSeparatorEngineWrapper(QObject):
    scanningChanged = Signal()
    processingChanged = Signal()
    scanCompleted = Signal("QVariant")
    progress = Signal("QVariant")
    processingCompleted = Signal("QVariant")
    error = Signal(str)

    # Exact extensions from One Tap V20.6.5
    PHOTO_EXTENSIONS = {
        # RAW formats
        '.cr2', '.cr3', '.crw', '.arw', '.nef', '.nrw', '.srf', '.sr2', '.dng',
        '.rw2', '.orf', '.gpr', '.pef', '.raf', '.rwl', '.srw', '.3fr', '.fff',
        '.iiq', '.k25', '.kdc', '.bay', '.erf', '.mef', '.mos', '.mrw', '.ptx',
        '.pxn', '.raw',
        # Raster & Web formats
        '.jpg', '.jpeg', '.jpe', '.png', '.heic', '.heif', '.hif', '.webp',
        '.tif', '.tiff', '.bmp', '.dib', '.gif', '.avif', '.jp2', '.j2k',
        '.jpf', '.jpx', '.jpm', '.mj2'
    }

    VIDEO_EXTENSIONS = {
        '.mov', '.mp4', '.mts', '.m4v', '.mkv', '.avi', '.flv', '.webm',
        '.wmv', '.3gp', '.mxf', '.mpg', '.mpeg', '.m2ts', '.ts', '.dv',
        '.r3d', '.braw', '.mpv', '.m2v', '.vob', '.ogv'
    }

    AUDIO_EXTENSIONS = {
        '.wav', '.mp3', '.aac', '.flac', '.ogg', '.wma', '.m4a', '.aif', '.aiff', '.opus'
    }

    PHOTO_SIDECAR_EXTENSIONS = {'.xmp', '.aae', '.nksc', '.cos'}
    VIDEO_SIDECAR_EXTENSIONS = {'.xml', '.srt', '.thm', '.vtt'}
    AUDIO_SIDECAR_EXTENSIONS = {'.pkf'}

    ALL_SIDECAR_EXTENSIONS = PHOTO_SIDECAR_EXTENSIONS | VIDEO_SIDECAR_EXTENSIONS | AUDIO_SIDECAR_EXTENSIONS
    SKIP_FILES = {'.ds_store', 'thumbs.db', 'desktop.ini'}

    def __init__(self, parent=None):
        super().__init__(parent)
        self._scanning = False
        self._processing = False
        self._active_scan_cache = None

    @Property(bool, notify=scanningChanged)
    def scanning(self):
        return self._scanning

    @Property(bool, notify=processingChanged)
    def isProcessing(self):
        return self._processing

    @classmethod
    def classify_file(cls, file_path, ext, stem_to_media_map):
        if ext in cls.PHOTO_EXTENSIONS:
            return 'photo'
        if ext in cls.VIDEO_EXTENSIONS:
            return 'video'
        if ext in cls.AUDIO_EXTENSIONS:
            return 'audio'

        if ext in cls.ALL_SIDECAR_EXTENSIONS:
            stem = Path(file_path).stem.lower()
            parent_type = stem_to_media_map.get(stem)
            if parent_type:
                return parent_type
            if ext in cls.PHOTO_SIDECAR_EXTENSIONS:
                return 'photo'
            if ext in cls.VIDEO_SIDECAR_EXTENSIONS:
                return 'video'
            if ext in cls.AUDIO_SIDECAR_EXTENSIONS:
                return 'audio'

        return 'misc'

    @Slot(str)
    def scanFolder(self, source_dir):
        if not source_dir:
            self.error.emit("Please select a valid source directory.")
            return

        clean_dir = os.path.abspath(os.path.normpath(source_dir.replace("file:///", "").strip()))
        if not os.path.isdir(clean_dir):
            self.error.emit(f"Source directory does not exist: {clean_dir}")
            return

        self._scanning = True
        self.scanningChanged.emit()

        def _worker():
            try:
                sub_dirs = []
                try:
                    with os.scandir(clean_dir) as it:
                        for entry in it:
                            if entry.is_dir() and not entry.name.startswith('.'):
                                sub_dirs.append(entry.name)
                except Exception as err:
                    self.error.emit(f"Failed to read directory: {err}")
                    return

                sub_dirs.sort(key=lambda s: [int(t) if t.isdigit() else t.lower() for t in re.split(r'(\d+)', s)])
                main_folder_name = os.path.basename(clean_dir)
                sanitized_main_folder_name = 'Source' if not main_folder_name or main_folder_name.endswith(':') else main_folder_name
                event_folders = sub_dirs if sub_dirs else [sanitized_main_folder_name]

                all_files = []
                photo_exts = {}
                video_exts = {}
                audio_exts = {}
                misc_exts = {}
                total_photos = 0
                total_videos = 0
                total_audio = 0
                total_misc = 0
                photo_bytes = 0

                for root, dirs, files in os.walk(clean_dir):
                    # Filter out hidden or junk dirs
                    dirs[:] = [d for d in dirs if not d.startswith('.')]

                    stem_map = {}
                    file_entries = []
                    for f in files:
                        lower_f = f.lower()
                        if lower_f.startswith('.') or lower_f in self.SKIP_FILES:
                            continue
                        file_entries.append(f)
                        ext = os.path.splitext(f)[1].lower()
                        stem = os.path.splitext(f)[0].lower()
                        if ext in self.PHOTO_EXTENSIONS:
                            stem_map[stem] = 'photo'
                        elif ext in self.VIDEO_EXTENSIONS:
                            stem_map[stem] = 'video'
                        elif ext in self.AUDIO_EXTENSIONS:
                            stem_map[stem] = 'audio'

                    for f in file_entries:
                        full_path = os.path.join(root, f)
                        ext = os.path.splitext(f)[1].lower()
                        ftype = self.classify_file(full_path, ext, stem_map)
                        try:
                            fsize = os.path.getsize(full_path)
                        except Exception:
                            fsize = 0

                        rel_from_source = os.path.relpath(full_path, clean_dir)
                        parts = rel_from_source.split(os.sep)
                        event_folder = parts[0] if len(parts) > 1 else sanitized_main_folder_name

                        finfo = {
                            "abs_path": full_path,
                            "event_folder": event_folder,
                            "rel_within_event": rel_from_source,
                            "ext": ext,
                            "type": ftype,
                            "size": fsize
                        }
                        all_files.append(finfo)

                        ext_display = ext if ext else '.no_ext'
                        if ftype == 'photo':
                            photo_exts[ext_display] = photo_exts.get(ext_display, 0) + 1
                            total_photos += 1
                            photo_bytes += fsize
                        elif ftype == 'video':
                            video_exts[ext_display] = video_exts.get(ext_display, 0) + 1
                            total_videos += 1
                        elif ftype == 'audio':
                            audio_exts[ext_display] = audio_exts.get(ext_display, 0) + 1
                            total_audio += 1
                        else:
                            misc_exts[ext_display] = misc_exts.get(ext_display, 0) + 1
                            total_misc += 1

                if not all_files:
                    self.error.emit("No files found in the source directory.")
                    return

                self._active_scan_cache = {
                    "source_dir": clean_dir,
                    "event_folders": event_folders,
                    "all_files": all_files,
                    "photo_bytes": photo_bytes
                }

                res = {
                    "total_files": len(all_files),
                    "total_photos": total_photos,
                    "total_videos": total_videos,
                    "total_audio": total_audio,
                    "total_misc": total_misc,
                    "photo_bytes": photo_bytes,
                    "photo_exts": photo_exts,
                    "video_exts": video_exts,
                    "audio_exts": audio_exts,
                    "misc_exts": misc_exts,
                    "event_folders": event_folders
                }
                self.scanCompleted.emit(res)
            except Exception as e:
                self.error.emit(str(e))
            finally:
                self._scanning = False
                self.scanningChanged.emit()

        threading.Thread(target=_worker, daemon=True).start()

    @Slot(str, str, str)
    def startProcessing(self, source_dir, dest_dir, action='copy'):
        clean_src = os.path.abspath(os.path.normpath(source_dir.replace("file:///", "").strip()))
        clean_dst = os.path.abspath(os.path.normpath(dest_dir.replace("file:///", "").strip()))

        if not self._active_scan_cache or self._active_scan_cache["source_dir"] != clean_src:
            # Auto-scan first then process
            self.error.emit("Please analyze the source directory before extracting.")
            return

        photo_files = [f for f in self._active_scan_cache["all_files"] if f["type"] == 'photo']
        total_photo_count = len(photo_files)

        if total_photo_count == 0:
            self.error.emit("No photo files found to process.")
            return

        self._processing = True
        self.processingChanged.emit()

        def _worker():
            total_bytes = sum(f["size"] for f in photo_files)
            photo_root = os.path.join(clean_dst, "Photo Data")
            raw_main_folder = os.path.basename(clean_src)
            main_folder_name = '' if not raw_main_folder or raw_main_folder.endswith(':') else raw_main_folder
            action_verb = "Copying" if action == "copy" else "Moving"

            copied_bytes = 0
            completed_count = 0
            error_count = 0
            start_time = time.time()
            last_progress_time = 0

            # Signal helper
            def send_progress(curr, tot, label="", log_entry=None, log_type="info", speed_mbps=0.0, eta_sec=0):
                data = {
                    "current": curr,
                    "total": tot,
                    "label": label,
                    "speed_mbps": speed_mbps,
                    "eta_sec": eta_sec,
                    "copied_bytes": copied_bytes,
                    "total_bytes": total_bytes,
                    "log_entry": log_entry,
                    "log_type": log_type
                }
                self.progress.emit(data)

            send_progress(0, total_photo_count, "Extracting photo data...", f"Started high-speed {action_verb}", "info")

            # Same volume check for atomic fast move
            src_drive = os.path.splitdrive(clean_src)[0].lower()
            dst_drive = os.path.splitdrive(clean_dst)[0].lower()
            is_same_volume = src_drive == dst_drive

            import concurrent.futures
            # 16-32 concurrency
            CONCURRENCY = 32 if (is_same_volume and action == 'move') else 16

            lock = threading.Lock()

            def transfer_single_file(finfo):
                nonlocal copied_bytes, completed_count, error_count, last_progress_time
                abs_path = finfo["abs_path"]
                rel_from_source = os.path.relpath(abs_path, clean_src)
                if main_folder_name:
                    photo_dest = os.path.join(photo_root, main_folder_name, rel_from_source)
                else:
                    photo_dest = os.path.join(photo_root, rel_from_source)

                dst_dir = os.path.dirname(photo_dest)
                os.makedirs(dst_dir, exist_ok=True)

                try:
                    if action == 'move':
                        if is_same_volume:
                            os.replace(abs_path, photo_dest)
                        else:
                            shutil.copy2(abs_path, photo_dest)
                            try:
                                os.unlink(abs_path)
                            except Exception:
                                pass
                    else:
                        shutil.copy2(abs_path, photo_dest)

                    with lock:
                        copied_bytes += finfo["size"]
                        completed_count += 1
                except Exception as err:
                    with lock:
                        error_count += 1
                    send_progress(completed_count, total_photo_count, f"{action_verb} photos...", f"{rel_from_source} -> {err}", "error")

                with lock:
                    now = time.time()
                    total_proc = completed_count + error_count
                    if now - last_progress_time > 0.05 or total_proc == total_photo_count:
                        last_progress_time = now
                        elapsed = max(0.001, now - start_time)
                        speed = round((copied_bytes / (1024 * 1024)) / elapsed, 1)
                        rem_bytes = max(0, total_bytes - copied_bytes)
                        bytes_sec = copied_bytes / elapsed if elapsed > 0 else 1
                        eta = int(rem_bytes / bytes_sec) if bytes_sec > 0 else 0

                        send_progress(total_proc, total_photo_count, f"{action_verb} photos... ({speed} MB/s)", None, "info", speed, eta)

            with concurrent.futures.ThreadPoolExecutor(max_workers=CONCURRENCY) as executor:
                list(executor.map(transfer_single_file, photo_files))

            # If move mode, cleanup empty folders
            if action == 'move':
                for root, dirs, files in os.walk(clean_src, topdown=False):
                    if root != clean_src:
                        try:
                            # If only junk remaining, remove them
                            remaining = os.listdir(root)
                            non_junk = [f for f in remaining if f.lower() not in self.SKIP_FILES and not f.startswith('.')]
                            if not non_junk:
                                for jf in remaining:
                                    try: os.unlink(os.path.join(root, jf))
                                    except Exception: pass
                                os.rmdir(root)
                        except Exception:
                            pass

            total_elapsed = max(0.001, time.time() - start_time)
            avg_speed = round((total_bytes / (1024 * 1024)) / total_elapsed, 1)
            done_msg = f"Complete! Processed {completed_count} photos in {total_elapsed:.1f}s (Avg {avg_speed} MB/s, {error_count} errors)"
            send_progress(total_photo_count, total_photo_count, "Complete!", done_msg, "success", avg_speed, 0)

            res = {
                "success_count": completed_count,
                "error_count": error_count,
                "elapsed_sec": round(total_elapsed, 1),
                "avg_speed_mbps": avg_speed,
                "dest_dir": photo_root
            }
            self.processingCompleted.emit(res)
            self._processing = False
            self.processingChanged.emit()

        threading.Thread(target=_worker, daemon=True).start()


class ThumbnailSeparatorEngineWrapper(QObject):
    scanningChanged = Signal()
    processingChanged = Signal()
    scanCompleted = Signal("QVariant")
    progress = Signal("QVariant")
    processingCompleted = Signal("QVariant")
    error = Signal(str)

    THUMBNAIL_EXTS = {'.jpg', '.jpeg', '.png', '.webp', '.bmp', '.tif', '.tiff'}

    def __init__(self, parent=None):
        super().__init__(parent)
        self._scanning = False
        self._processing = False
        self._active_scan_cache = None

    @Property(bool, notify=scanningChanged)
    def scanning(self):
        return self._scanning

    @Property(bool, notify=processingChanged)
    def isProcessing(self):
        return self._processing

    @Slot(str)
    def scan(self, source_dir):
        self.scanFolder(source_dir)

    @Slot(str)
    def scanFolder(self, source_dir):
        if not source_dir:
            self.error.emit("Please select a valid source directory.")
            return

        clean_path = os.path.abspath(os.path.normpath(source_dir.replace("file:///", "").strip()))
        if not os.path.isdir(clean_path):
            self.error.emit(f"Source directory does not exist: {clean_path}")
            return

        self._scanning = True
        self.scanningChanged.emit()

        def _worker():
            try:
                items = []
                p_count = 0
                v_count = 0
                total_bytes = 0

                for root_dir, dirs, files in os.walk(clean_path):
                    for f in files:
                        ext = os.path.splitext(f)[1].lower()
                        if ext in self.THUMBNAIL_EXTS:
                            stem = os.path.splitext(f)[0]
                            # Match if stem ends with _p or _v (case-insensitive)
                            is_p = bool(re.search(r'_[pP]$', stem))
                            is_v = bool(re.search(r'_[vV]$', stem))

                            if is_p or is_v:
                                full_p = os.path.join(root_dir, f)
                                rel_p = os.path.relpath(full_p, clean_path)
                                rel_d = os.path.dirname(rel_p)
                                if rel_d == ".":
                                    rel_d = ""

                                try:
                                    sz = os.path.getsize(full_p)
                                except Exception:
                                    sz = 0

                                total_bytes += sz
                                if is_p:
                                    p_count += 1
                                    t_type = "photo"
                                    t_label = "_P Photo Thumb"
                                else:
                                    v_count += 1
                                    t_type = "video"
                                    t_label = "_V Video Thumb"

                                items.append({
                                    "filename": f,
                                    "stem": stem,
                                    "ext": ext,
                                    "type": t_type,
                                    "typeLabel": t_label,
                                    "sourcePath": full_p,
                                    "relPath": rel_p,
                                    "relDir": rel_d,
                                    "sizeBytes": sz,
                                    "status": "READY"
                                })

                res = {
                    "sourceDir": clean_path,
                    "totalThumbnails": len(items),
                    "pCount": p_count,
                    "vCount": v_count,
                    "totalBytes": total_bytes,
                    "items": items
                }
                self._active_scan_cache = res
                self.scanCompleted.emit(res)
            except Exception as e:
                self.error.emit(f"Scan failed: {str(e)}")
            finally:
                self._scanning = False
                self.scanningChanged.emit()

        threading.Thread(target=_worker, daemon=True).start()

    @Slot(str, str, str, "QVariant")
    def executeTransfer(self, source_dir, dest_dir, action_mode="copy", selected_items=None):
        if not source_dir or not dest_dir:
            self.error.emit("Source and destination directories are required.")
            return

        clean_src = os.path.abspath(os.path.normpath(source_dir.replace("file:///", "").strip()))
        clean_dest = os.path.abspath(os.path.normpath(dest_dir.replace("file:///", "").strip()))

        if not os.path.isdir(clean_src):
            self.error.emit(f"Source directory does not exist: {clean_src}")
            return

        os.makedirs(clean_dest, exist_ok=True)

        selected_items = to_py_variant(selected_items)
        if selected_items is None and self._active_scan_cache:
            selected_items = self._active_scan_cache.get("items", [])

        if not selected_items:
            self.error.emit("No thumbnail files selected to process.")
            return

        self._processing = True
        self.processingChanged.emit()

        def _worker():
            total = len(selected_items)
            done = 0
            success_count = 0
            error_count = 0
            total_bytes_transferred = 0
            start_time = time.time()
            transferred_records = []
            errors = []

            mode = action_mode.lower().strip()

            for item in selected_items:
                src_path = item.get("sourcePath") or os.path.join(clean_src, item.get("relPath", ""))
                rel_dir = item.get("relDir", "")
                filename = item.get("filename", os.path.basename(src_path))

                target_dir = os.path.join(clean_dest, rel_dir) if rel_dir else clean_dest
                target_path = os.path.join(target_dir, filename)

                try:
                    if not os.path.exists(src_path):
                        raise FileNotFoundError(f"Source file not found: {src_path}")

                    os.makedirs(target_dir, exist_ok=True)
                    f_size = os.path.getsize(src_path)

                    if mode == "move":
                        shutil.move(src_path, target_path)
                    else:
                        shutil.copy2(src_path, target_path)

                    total_bytes_transferred += f_size
                    success_count += 1

                    transferred_records.append({
                        "filename": filename,
                        "type": item.get("type", ""),
                        "sourcePath": src_path,
                        "targetPath": target_path,
                        "relDir": rel_dir,
                        "mode": mode.upper(),
                        "status": "SUCCESS"
                    })
                except Exception as ex:
                    error_count += 1
                    err_msg = str(ex)
                    errors.append({"file": filename, "error": err_msg})
                    transferred_records.append({
                        "filename": filename,
                        "type": item.get("type", ""),
                        "sourcePath": src_path,
                        "targetPath": target_path,
                        "relDir": rel_dir,
                        "mode": mode.upper(),
                        "status": f"ERROR: {err_msg}"
                    })

                done += 1
                elapsed = time.time() - start_time
                speed_bps = total_bytes_transferred / elapsed if elapsed > 0 else 0
                speed_mbps = round(speed_bps / (1024 * 1024), 2)
                eta_sec = int((total - done) / (done / elapsed)) if done > 0 and elapsed > 0 else 0

                self.progress.emit({
                    "current": done,
                    "total": total,
                    "current_file": filename,
                    "speed_mbps": speed_mbps,
                    "eta_sec": eta_sec,
                    "log_entry": f"[{mode.upper()}] {filename} -> {rel_dir or '.'}",
                    "log_type": "success" if not errors else "info"
                })

            duration_sec = round(time.time() - start_time, 2)
            avg_speed_mbps = round((total_bytes_transferred / max(duration_sec, 0.001)) / (1024 * 1024), 2)

            # Auto-export CSV Report to Document folder inside dest_dir
            doc_dir = os.path.join(clean_dest, "Document")
            os.makedirs(doc_dir, exist_ok=True)
            ts = time.strftime("%Y%m%d_%H%M%S")
            csv_path = os.path.join(doc_dir, f"Thumbnail_Separation_Report_{ts}.csv")

            try:
                with open(csv_path, "w", newline="", encoding="utf-8-sig") as f_csv:
                    writer = csv.writer(f_csv)
                    writer.writerow(["SR NO", "FILENAME", "TYPE", "MODE", "REL DIRECTORY", "TARGET PATH", "STATUS", "TIMESTAMP"])
                    for idx, rec in enumerate(transferred_records, 1):
                        writer.writerow([
                            idx,
                            rec.get("filename", ""),
                            rec.get("type", ""),
                            rec.get("mode", ""),
                            rec.get("relDir", ""),
                            rec.get("targetPath", ""),
                            rec.get("status", ""),
                            time.strftime("%Y-%m-%d %H:%M:%S")
                        ])
            except Exception as e_csv:
                print(f"Thumbnail report CSV error: {e_csv}")

            res = {
                "success_count": success_count,
                "error_count": error_count,
                "total_bytes": total_bytes_transferred,
                "elapsed_sec": duration_sec,
                "avg_speed_mbps": avg_speed_mbps,
                "report_csv": csv_path,
                "dest_dir": clean_dest,
                "errors": errors
            }

            self.processingCompleted.emit(res)
            self._processing = False
            self.processingChanged.emit()

        threading.Thread(target=_worker, daemon=True).start()


class RemainingPhotosCollectorEngineWrapper(QObject):
    scanningChanged = Signal()
    shiftingChanged = Signal()
    undoingChanged = Signal()
    scanCompleted = Signal(dict)
    shiftProgress = Signal(int, int, int, str)
    shiftCompleted = Signal(dict)
    undoCompleted = Signal(int, int)
    error = Signal(str)
    currentFolderChanged = Signal()
    shiftPercentChanged = Signal()
    progressStrChanged = Signal()

    IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.webp', '.bmp', '.tiff', '.tif', '.cr2', '.cr3', '.nef', '.arw', '.raf', '.dng', '.orf', '.rw2', '.pef', '.srw', '.heic'}

    def __init__(self, parent=None):
        super().__init__(parent)
        self._scanning = False
        self._shifting = False
        self._undoing = False
        self._currentFolder = ""
        self._shiftPercent = 0
        self._progressStr = ""

    @Property(bool, notify=scanningChanged)
    def scanning(self): return self._scanning

    @Property(bool, notify=shiftingChanged)
    def shifting(self): return self._shifting

    @Property(bool, notify=undoingChanged)
    def undoing(self): return self._undoing

    @Property(str, notify=currentFolderChanged)
    def currentFolder(self): return self._currentFolder

    @Property(int, notify=shiftPercentChanged)
    def shiftPercent(self): return self._shiftPercent

    @Property(str, notify=progressStrChanged)
    def progressStr(self): return self._progressStr

    @Slot(str)
    def scan(self, source_dir):
        self.scanFolders(source_dir)

    @Slot(str)
    def scanFolders(self, source_dir):
        if not source_dir:
            self.error.emit("Please select a valid main source folder.")
            return

        clean_path = os.path.abspath(os.path.normpath(source_dir.replace("file:///", "").strip()))
        if not os.path.isdir(clean_path):
            self.error.emit(f"Directory does not exist: {clean_path}")
            return

        self._scanning = True
        self.scanningChanged.emit()

        def _scan_worker():
            found_items = []
            date_folders_set = set()
            total_photos_count = 0
            total_bytes = 0
            processed_roots = set()

            def is_remaining_folder(name):
                n = name.lower().strip()
                if 'remain' in n:
                    return True
                if re.search(r'(?:^|[\s_\-])rem(?:aining)?(?:[\s_\-]|photos?|$)', n):
                    return True
                return False

            def extract_date_folder_name(full_path, base_root):
                rel = os.path.relpath(full_path, base_root)
                parts = [p for p in rel.replace('\\', '/').split('/') if p and p != '.']
                for part in parts[:-1]:
                    if re.search(r'\d{4}[-_.]\d{2}[-_.]\d{2}', part) or re.search(r'\d{2}[-_.]\d{2}[-_.]\d{4}', part) or re.match(r'^(?:day|date|event|session)[\s_\-]*\d+', part, re.I):
                        return part
                if len(parts) > 1:
                    return parts[0]
                return os.path.basename(os.path.dirname(full_path)) or os.path.basename(base_root)

            try:
                for root, dirs, files in os.walk(clean_path):
                    # Skip if inside an already processed remaining folder
                    if any(root.startswith(pr + os.sep) for pr in processed_roots):
                        continue

                    folder_name = os.path.basename(root).strip()
                    if is_remaining_folder(folder_name):
                        processed_roots.add(root)
                        parent_date_folder = extract_date_folder_name(root, clean_path)
                        date_folders_set.add(parent_date_folder)

                        photo_count = 0
                        folder_size = 0
                        for sub_root, _, sub_files in os.walk(root):
                            for f in sub_files:
                                ext = os.path.splitext(f)[1].lower()
                                if ext in self.IMAGE_EXTS:
                                    photo_count += 1
                                fp = os.path.join(sub_root, f)
                                try:
                                    folder_size += os.path.getsize(fp)
                                except Exception:
                                    pass

                        total_photos_count += photo_count
                        total_bytes += folder_size

                        if folder_size > 1024 * 1024 * 1024:
                            size_str = f"{folder_size / (1024**3):.2f} GB"
                        else:
                            size_str = f"{folder_size / (1024**2):.1f} MB"

                        rel = os.path.relpath(root, clean_path)

                        found_items.append({
                            "date": parent_date_folder,
                            "dateFolder": parent_date_folder,
                            "dateFolderName": parent_date_folder,
                            "folderName": folder_name,
                            "relPath": rel,
                            "fullPath": root,
                            "photoCount": photo_count,
                            "fileCount": photo_count,
                            "files": photo_count,
                            "sizeBytes": folder_size,
                            "sizeStr": size_str,
                            "size": size_str,
                            "status": "Ready"
                        })

                if total_bytes > 1024 * 1024 * 1024:
                    total_size_str = f"{total_bytes / (1024**3):.2f} GB"
                else:
                    total_size_str = f"{total_bytes / (1024**2):.1f} MB"

                result = {
                    "success": True,
                    "items": found_items,
                    "folders": found_items,
                    "manifest": found_items,
                    "totalFolders": len(found_items),
                    "total_folders": len(found_items),
                    "totalDateFolders": len(date_folders_set),
                    "totalPhotos": total_photos_count,
                    "total_photos": total_photos_count,
                    "totalSizeBytes": total_bytes,
                    "totalSizeStr": total_size_str,
                    "sourceDir": clean_path
                }
                self._last_scanned_items = found_items
                self.scanCompleted.emit(result)
            except Exception as e:
                self.error.emit(f"Scan failed: {str(e)}")
            finally:
                self._scanning = False
                self.scanningChanged.emit()

        threading.Thread(target=_scan_worker, daemon=True).start()

    @Slot(str, str)
    def shift(self, source_dir, dest_dir):
        self.shift(source_dir, dest_dir, "move")

    @Slot(str, str, str)
    def shift(self, source_dir, dest_dir, mode="move"):
        items = getattr(self, '_last_scanned_items', [])
        self.executeShift({
            "sourceDir": source_dir,
            "targetDir": dest_dir,
            "mode": mode,
            "items": items
        })

    @Slot("QVariant")
    def executeShift(self, payload):
        payload = to_py_variant(payload)
        source_dir = payload.get("sourceDir", "") if isinstance(payload, dict) else ""
        target_dir = payload.get("targetDir", "") if isinstance(payload, dict) else ""
        mode = payload.get("mode", "move") if isinstance(payload, dict) else "move"
        items = payload.get("items", []) if isinstance(payload, dict) else []

        if not items and hasattr(self, '_last_scanned_items') and self._last_scanned_items:
            items = self._last_scanned_items

        if not target_dir:
            self.error.emit("Please select a target destination folder.")
            return

        tgt_clean = os.path.abspath(os.path.normpath(target_dir.replace("file:///", "").strip()))
        os.makedirs(tgt_clean, exist_ok=True)

        self._shifting = True
        self.shiftingChanged.emit()

        def _shift_worker():
            succ = 0
            fail = 0
            shifted_manifest_items = []
            total = len(items)

            try:
                for i, itm in enumerate(items):
                    src_full = itm.get("fullPath", "")
                    date_folder = itm.get("dateFolder", "Remaining Photos")
                    photo_count = itm.get("photoCount", 0)
                    size_bytes = itm.get("sizeBytes", 0)
                    size_str = itm.get("sizeStr", "0 MB")

                    self._currentFolder = f"{date_folder} ({src_full})"
                    self._shiftPercent = int(((i + 1) / max(total, 1)) * 100)
                    self._progressStr = f"{i + 1} of {total} folders"
                    self.currentFolderChanged.emit()
                    self.shiftPercentChanged.emit()
                    self.progressStrChanged.emit()
                    self.shiftProgress.emit(i + 1, total, self._shiftPercent, self._currentFolder)

                    if not os.path.exists(src_full):
                        fail += 1
                        shifted_manifest_items.append({
                            "originalPath": src_full,
                            "newPath": "",
                            "dateFolder": date_folder,
                            "photoCount": photo_count,
                            "sizeBytes": size_bytes,
                            "sizeStr": size_str,
                            "status": "Failed",
                            "error": "Source folder not found"
                        })
                        continue

                    # Target directory named after the Date folder
                    dest_dir = os.path.join(tgt_clean, date_folder)
                    
                    # If destination already exists, resolve collision
                    final_dest = dest_dir
                    if os.path.exists(final_dest) and os.path.abspath(final_dest) != os.path.abspath(src_full):
                        counter = 1
                        while os.path.exists(final_dest):
                            final_dest = f"{dest_dir}_{counter}"
                            counter += 1

                    try:
                        if mode == "move":
                            shutil.move(src_full, final_dest)
                        else:
                            shutil.copytree(src_full, final_dest, dirs_exist_ok=True)

                        succ += 1
                        shifted_manifest_items.append({
                            "originalPath": src_full,
                            "newPath": final_dest,
                            "dateFolder": date_folder,
                            "photoCount": photo_count,
                            "sizeBytes": size_bytes,
                            "sizeStr": size_str,
                            "status": "Success",
                            "error": ""
                        })
                    except Exception as ex:
                        print(f"Error shifting {src_full}: {ex}")
                        fail += 1
                        shifted_manifest_items.append({
                            "originalPath": src_full,
                            "newPath": final_dest,
                            "dateFolder": date_folder,
                            "photoCount": photo_count,
                            "sizeBytes": size_bytes,
                            "sizeStr": size_str,
                            "status": "Failed",
                            "error": str(ex)
                        })

                # 1. Create document folder and excel report
                doc_dir, excel_report_path = self._create_excel_report(
                    tgt_clean, source_dir, mode, shifted_manifest_items, total, succ, fail
                )

                # 2. Save JSON Manifest
                manifest_path = os.path.join(tgt_clean, "remaining_shift_manifest.json")
                try:
                    manifest_data = {
                        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
                        "sourceDir": source_dir,
                        "targetDir": tgt_clean,
                        "documentFolder": doc_dir,
                        "excelReportPath": excel_report_path,
                        "mode": mode,
                        "totalItems": len(shifted_manifest_items),
                        "items": shifted_manifest_items
                    }
                    with open(manifest_path, "w", encoding="utf-8") as f:
                        json.dump(manifest_data, f, indent=2)
                except Exception:
                    manifest_path = ""

                self.shiftCompleted.emit({
                    "success": True,
                    "successful": succ,
                    "failed": fail,
                    "total": total,
                    "manifestPath": manifest_path,
                    "documentFolder": doc_dir,
                    "excelReportPath": excel_report_path,
                    "shifted_folders": succ,
                    "shifted_photos": sum(itm.get("photoCount", 0) for itm in shifted_manifest_items if itm.get("status") == "Success")
                })
            except Exception as e:
                self.error.emit(f"Shift execution failed: {str(e)}")
            finally:
                self._shifting = False
                self.shiftingChanged.emit()

        threading.Thread(target=_shift_worker, daemon=True).start()
        # 1. Resolve or create 'document' folder inside target directory
        doc_dir = None
        if os.path.exists(target_dir):
            for name in os.listdir(target_dir):
                if name.lower() in {"document", "documents"} and os.path.isdir(os.path.join(target_dir, name)):
                    doc_dir = os.path.join(target_dir, name)
                    break
        if not doc_dir:
            doc_dir = os.path.join(target_dir, "document")
            os.makedirs(doc_dir, exist_ok=True)

        timestamp_str = time.strftime("%Y%m%d_%H%M%S")
        excel_filename = f"Remaining_Shift_Report_{timestamp_str}.xlsx"
        excel_path = os.path.join(doc_dir, excel_filename)

        total_photos = sum(itm.get("photoCount", 0) for itm in shifted_items if itm.get("status") == "Success")
        total_bytes = sum(itm.get("sizeBytes", 0) for itm in shifted_items if itm.get("status") == "Success")
        if total_bytes > 1024 * 1024 * 1024:
            total_size_str = f"{total_bytes / (1024**3):.2f} GB"
        else:
            total_size_str = f"{total_bytes / (1024**2):.1f} MB"

        now_str = time.strftime("%Y-%m-%d %H:%M:%S")

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()

            # --- Sheet 1: Summary ---
            ws_summary = wb.active
            ws_summary.title = "Summary"
            ws_summary.views.sheetView[0].showGridLines = True

            HEADER_FILL = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")
            SUBHEADER_FILL = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
            LABEL_FILL = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
            SUCCESS_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            DANGER_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

            thin_border = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )

            # Title Block
            ws_summary.merge_cells("A1:D1")
            title_cell = ws_summary["A1"]
            title_cell.value = "SEQUORA STUDIO — REMAINING PHOTOS SHIFT REPORT"
            title_cell.font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
            title_cell.fill = HEADER_FILL
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_summary.row_dimensions[1].height = 36

            # Subtitle Block
            ws_summary.merge_cells("A2:D2")
            sub_cell = ws_summary["A2"]
            sub_cell.value = f"Generated on: {now_str}  |  Operation Mode: {mode.upper()}"
            sub_cell.font = Font(name="Segoe UI", size=10, italic=True, color="FFFFFF")
            sub_cell.fill = SUBHEADER_FILL
            sub_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_summary.row_dimensions[2].height = 22

            summary_rows = [
                ("Execution Timestamp", now_str),
                ("Shift Mode", mode.upper()),
                ("Source Directory", source_dir),
                ("Target Destination", target_dir),
                ("Report Location", excel_path),
                ("Total Folders Processed", total_items),
                ("Successfully Shifted", succ_count),
                ("Failed Folders", fail_count),
                ("Total Photos Shifted", total_photos),
                ("Total Data Shifted", total_size_str)
            ]

            row_idx = 4
            for label, val in summary_rows:
                cell_lbl = ws_summary.cell(row=row_idx, column=1, value=label)
                cell_lbl.font = Font(name="Segoe UI", size=11, bold=True, color="1E293B")
                cell_lbl.fill = LABEL_FILL
                cell_lbl.border = thin_border
                cell_lbl.alignment = Alignment(vertical="center")

                ws_summary.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
                cell_val = ws_summary.cell(row=row_idx, column=2, value=val)
                cell_val.font = Font(name="Segoe UI", size=11, color="0F172A")
                cell_val.alignment = Alignment(vertical="center")

                for col in range(2, 5):
                    ws_summary.cell(row=row_idx, column=col).border = thin_border

                if label == "Successfully Shifted":
                    cell_val.fill = SUCCESS_FILL
                    cell_val.font = Font(name="Segoe UI", size=11, bold=True, color="166534")
                elif label == "Failed Folders" and fail_count > 0:
                    cell_val.fill = DANGER_FILL
                    cell_val.font = Font(name="Segoe UI", size=11, bold=True, color="991B1B")

                ws_summary.row_dimensions[row_idx].height = 24
                row_idx += 1

            ws_summary.column_dimensions['A'].width = 26
            ws_summary.column_dimensions['B'].width = 30
            ws_summary.column_dimensions['C'].width = 30
            ws_summary.column_dimensions['D'].width = 30

            # --- Sheet 2: Folder Details ---
            ws_details = wb.create_sheet(title="Folder Details")
            ws_details.views.sheetView[0].showGridLines = True

            headers = [
                "S.No", "Date / Folder Name", "Photos", "Size", "Mode", "Status", "Original Source Path", "Shifted Target Path", "Notes / Error"
            ]

            ws_details.row_dimensions[1].height = 28
            for col_idx, h_text in enumerate(headers, 1):
                c = ws_details.cell(row=1, column=col_idx, value=h_text)
                c.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                c.fill = HEADER_FILL
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border

            for item_idx, itm in enumerate(shifted_items, 1):
                r = item_idx + 1
                ws_details.row_dimensions[r].height = 22

                is_succ = itm.get("status") == "Success"
                status_text = "SUCCESS" if is_succ else "FAILED"
                status_fill = SUCCESS_FILL if is_succ else DANGER_FILL
                status_font_color = "166534" if is_succ else "991B1B"

                row_data = [
                    (item_idx, "center"),
                    (itm.get("dateFolder", ""), "left"),
                    (itm.get("photoCount", 0), "center"),
                    (itm.get("sizeStr", ""), "center"),
                    (mode.upper(), "center"),
                    (status_text, "center"),
                    (itm.get("originalPath", ""), "left"),
                    (itm.get("newPath", ""), "left"),
                    (itm.get("error", ""), "left")
                ]

                for col_idx, (val, align) in enumerate(row_data, 1):
                    c = ws_details.cell(row=r, column=col_idx, value=val)
                    c.font = Font(name="Segoe UI", size=10)
                    c.border = thin_border
                    c.alignment = Alignment(horizontal=align, vertical="center")
                    if col_idx == 6:  # Status
                        c.fill = status_fill
                        c.font = Font(name="Segoe UI", size=10, bold=True, color=status_font_color)

            col_widths = {1: 8, 2: 24, 3: 12, 4: 14, 5: 12, 6: 14, 7: 45, 8: 45, 9: 30}
            for col_idx, width in col_widths.items():
                col_letter = get_column_letter(col_idx)
                ws_details.column_dimensions[col_letter].width = width

            wb.save(excel_path)

            # Keep a convenience copy named Remaining_Shift_Report_Latest.xlsx
            latest_path = os.path.join(doc_dir, "Remaining_Shift_Report_Latest.xlsx")
            try:
                shutil.copyfile(excel_path, latest_path)
            except Exception:
                pass

        except Exception as e:
            print(f"Error generating Excel report: {e}")
            try:
                csv_path = os.path.join(doc_dir, f"Remaining_Shift_Report_{timestamp_str}.csv")
                with open(csv_path, "w", encoding="utf-8") as cf:
                    cf.write("S.No,Date Folder,Photos,Size,Mode,Status,Original Path,New Path,Error\n")
                    for idx, itm in enumerate(shifted_items, 1):
                        cf.write(f'{idx},"{itm.get("dateFolder","")}",{itm.get("photoCount",0)},"{itm.get("sizeStr","")}","{mode.upper()}","{itm.get("status","")}","{itm.get("originalPath","")}","{itm.get("newPath","")}","{itm.get("error","")}"\n')
                excel_path = csv_path
            except Exception:
                pass

        return doc_dir, excel_path

    @Slot("QVariant")
    def executeShift(self, payload):
        payload = to_py_variant(payload)
        source_dir = payload.get("sourceDir", "") if isinstance(payload, dict) else ""
        target_dir = payload.get("targetDir", "")
        mode = payload.get("mode", "move")
        items = payload.get("items", [])

        if not target_dir:
            self.error.emit("Please select a target destination folder.")
            return

        tgt_clean = os.path.abspath(os.path.normpath(target_dir.replace("file:///", "").strip()))
        os.makedirs(tgt_clean, exist_ok=True)

        self._shifting = True
        self.shiftingChanged.emit()
        QApplication.processEvents()

        succ = 0
        fail = 0
        shifted_manifest_items = []
        total = len(items)

        try:
            for i, itm in enumerate(items):
                src_full = itm.get("fullPath", "")
                date_folder = itm.get("dateFolder", "Remaining Photos")
                photo_count = itm.get("photoCount", 0)
                size_bytes = itm.get("sizeBytes", 0)
                size_str = itm.get("sizeStr", "0 MB")

                self._currentFolder = f"{date_folder} ({src_full})"
                self._shiftPercent = int(((i + 1) / max(total, 1)) * 100)
                self._progressStr = f"{i + 1} of {total} folders"
                self.currentFolderChanged.emit()
                self.shiftPercentChanged.emit()
                self.progressStrChanged.emit()
                self.shiftProgress.emit(i + 1, total, self._shiftPercent, self._currentFolder)

                if not os.path.exists(src_full):
                    fail += 1
                    shifted_manifest_items.append({
                        "originalPath": src_full,
                        "newPath": "",
                        "dateFolder": date_folder,
                        "photoCount": photo_count,
                        "sizeBytes": size_bytes,
                        "sizeStr": size_str,
                        "status": "Failed",
                        "error": "Source folder not found"
                    })
                    continue

                # Target directory named after the Date folder
                dest_dir = os.path.join(tgt_clean, date_folder)
                
                # If destination already exists, resolve collision
                final_dest = dest_dir
                if os.path.exists(final_dest) and os.path.abspath(final_dest) != os.path.abspath(src_full):
                    counter = 1
                    while os.path.exists(final_dest):
                        final_dest = f"{dest_dir}_{counter}"
                        counter += 1

                try:
                    if mode == "move":
                        shutil.move(src_full, final_dest)
                    else:
                        shutil.copytree(src_full, final_dest, dirs_exist_ok=True)

                    succ += 1
                    shifted_manifest_items.append({
                        "originalPath": src_full,
                        "newPath": final_dest,
                        "dateFolder": date_folder,
                        "photoCount": photo_count,
                        "sizeBytes": size_bytes,
                        "sizeStr": size_str,
                        "status": "Success",
                        "error": ""
                    })
                except Exception as ex:
                    print(f"Error shifting {src_full}: {ex}")
                    fail += 1
                    shifted_manifest_items.append({
                        "originalPath": src_full,
                        "newPath": final_dest,
                        "dateFolder": date_folder,
                        "photoCount": photo_count,
                        "sizeBytes": size_bytes,
                        "sizeStr": size_str,
                        "status": "Failed",
                        "error": str(ex)
                    })

            # 1. Create document folder and excel report
            doc_dir, excel_report_path = self._create_excel_report(
                tgt_clean, source_dir, mode, shifted_manifest_items, total, succ, fail
            )

            # 2. Save JSON Manifest
            manifest_path = os.path.join(tgt_clean, "remaining_shift_manifest.json")
            try:
                manifest_data = {
                    "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
                    "sourceDir": source_dir,
                    "targetDir": tgt_clean,
                    "documentFolder": doc_dir,
                    "excelReportPath": excel_report_path,
                    "mode": mode,
                    "totalItems": len(shifted_manifest_items),
                    "items": shifted_manifest_items
                }
                with open(manifest_path, "w", encoding="utf-8") as f:
                    json.dump(manifest_data, f, indent=2)
            except Exception:
                manifest_path = ""

            self.shiftCompleted.emit({
                "success": True,
                "successful": succ,
                "failed": fail,
                "total": total,
                "manifestPath": manifest_path,
                "documentFolder": doc_dir,
                "excelReportPath": excel_report_path
            })
        except Exception as e:
            self.error.emit(f"Shift execution failed: {str(e)}")
        finally:
            self._shifting = False
            self.shiftingChanged.emit()

    @Slot(str)
    def undoShift(self, manifest_path):
        if not manifest_path or not os.path.exists(manifest_path):
            self.error.emit("Shift manifest file not found.")
            return

        self._undoing = True
        self.undoingChanged.emit()

        restored = 0
        failed = 0
        try:
            with open(manifest_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            if data.get("mode") != "move":
                self.error.emit("Undo is only available for 'move' mode.")
                return

            for itm in data.get("items", []):
                orig = itm.get("originalPath", "")
                new_p = itm.get("newPath", "")
                if new_p and orig and os.path.exists(new_p):
                    os.makedirs(os.path.dirname(orig), exist_ok=True)
                    shutil.move(new_p, orig)
                    restored += 1
                else:
                    failed += 1

            self.undoCompleted.emit(restored, failed)
        except Exception as e:
            self.error.emit(f"Undo failed: {str(e)}")
        finally:
            self._undoing = False
            self.undoingChanged.emit()


class NativeDialogsWrapper(QObject):
    def __init__(self, parent=None):
        super().__init__(parent)

    @Slot(str, result=str)
    @Slot(str, str, result=str)
    def selectDirectory(self, title="Select Directory", start_dir=""):
        path = QFileDialog.getExistingDirectory(None, title, start_dir or "")
        return path or ""

    @Slot(str, result=str)
    @Slot(str, str, result=str)
    def chooseDirectory(self, title="Select Directory", start_dir=""):
        path = QFileDialog.getExistingDirectory(None, title, start_dir or "")
        return path or ""

    @Slot(str, str, result=str)
    def selectOpenFile(self, title, filter_str):
        path, _ = QFileDialog.getOpenFileName(None, title, "", filter_str)
        return path or ""

    @Slot(str, str, result="QVariantList")
    def selectOpenFiles(self, title, filter_str):
        paths, _ = QFileDialog.getOpenFileNames(None, title, "", filter_str)
        return paths or []

    @Slot(str, str, result=str)
    def selectSaveFile(self, title, filter_str):
        path, _ = QFileDialog.getSaveFileName(None, title, "", filter_str)
        return path or ""

    @Slot(str)
    def openPath(self, target_path):
        if not target_path:
            return
        clean = os.path.abspath(os.path.normpath(target_path.replace("file:///", "").strip()))
        if os.path.exists(clean):
            try:
                if sys.platform == "win32":
                    os.startfile(clean)
                elif sys.platform == "darwin":
                    import subprocess
                    subprocess.run(["open", clean])
                else:
                    import subprocess
                    subprocess.run(["xdg-open", clean])
            except Exception as e:
                print(f"Error opening path {clean}: {e}")

    @Slot(str)
    def openFolder(self, target_path):
        self.openPath(target_path)

    @Slot(str)
    def showInFolder(self, target_path):
        """Highlights the downloaded file or selected directory in Windows File Explorer."""
        if not target_path:
            return
        clean = os.path.abspath(os.path.normpath(target_path.replace("file:///", "").strip()))
        if os.path.exists(clean):
            try:
                if sys.platform == "win32":
                    import subprocess
                    subprocess.Popen(f'explorer /select,"{clean}"')
                elif sys.platform == "darwin":
                    import subprocess
                    subprocess.run(["open", "-R", clean])
                else:
                    self.openFolder(os.path.dirname(clean))
            except Exception as e:
                print(f"Error revealing path {clean}: {e}")
        else:
            # If specific file is missing, open parent folder
            parent_dir = os.path.dirname(clean)
            if os.path.exists(parent_dir):
                self.openFolder(parent_dir)

    @Slot(result=str)
    def getDefaultDownloadPath(self):
        """Returns standard system Downloads folder."""
        downloads_dir = str(Path.home() / "Downloads")
        os.makedirs(downloads_dir, exist_ok=True)
        return downloads_dir.replace("\\", "/")



def copy_cell_style(src_cell, dst_cell):
    """Deep-copies all visual styling attributes from src_cell to dst_cell."""
    if not src_cell.has_style:
        return
    import copy
    if src_cell.font:
        dst_cell.font = copy.copy(src_cell.font)
    if src_cell.fill:
        dst_cell.fill = copy.copy(src_cell.fill)
    if src_cell.border:
        dst_cell.border = copy.copy(src_cell.border)
    if src_cell.alignment:
        dst_cell.alignment = copy.copy(src_cell.alignment)
    if src_cell.number_format:
        dst_cell.number_format = src_cell.number_format
    if src_cell.protection:
        dst_cell.protection = copy.copy(src_cell.protection)


def copy_col_dimensions(ws_src, ws_dst, src_col_idx, dst_col_idx):
    """Copies column width from source worksheet to destination worksheet."""
    from openpyxl.utils import get_column_letter
    src_letter = get_column_letter(src_col_idx)
    dst_letter = get_column_letter(dst_col_idx)
    if src_letter in ws_src.column_dimensions:
        src_dim = ws_src.column_dimensions[src_letter]
        if src_dim.width is not None:
            ws_dst.column_dimensions[dst_letter].width = src_dim.width
        if src_dim.hidden:
            ws_dst.column_dimensions[dst_letter].hidden = True


def merge_inventory_reports(file_paths, output_path=None):
    """
    Merges multiple Excel inventory report files into a single workbook with full style fidelity.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not file_paths or len(file_paths) < 1:
        raise ValueError("At least one Excel file path must be provided for merging.")

    valid_files = [str(f) for f in file_paths if os.path.isfile(f) and f.lower().endswith(('.xlsx', '.xlsm'))]
    if not valid_files:
        raise FileNotFoundError(f"No valid Excel files found in {file_paths}")

    if output_path is None:
        first_dir = os.path.dirname(os.path.abspath(valid_files[0]))
        first_stem = Path(valid_files[0]).stem
        # Extract base prefix (e.g. '1448-03-12' from '1448-03-12 _01M_Inventory_Report')
        clean_prefix = re.sub(r'(_01M|_02E)?_?Inventory_Report.*$', '', first_stem, flags=re.IGNORECASE).rstrip(' _')
        if not clean_prefix:
            clean_prefix = first_stem
        merged_filename = f"{clean_prefix}_Inventory_Report_MERGED.xlsx"
        output_path = os.path.join(first_dir, merged_filename)

    loaded_wbs = []
    for fp in valid_files:
        wb = openpyxl.load_workbook(fp)
        loaded_wbs.append((fp, wb))

    wb_out = openpyxl.Workbook()
    if wb_out.sheetnames:
        wb_out.remove(wb_out.active)

    sheet_order = []
    for _, wb in loaded_wbs:
        for sname in wb.sheetnames:
            if sname not in sheet_order:
                sheet_order.append(sname)

    stats = {
        "files_merged": len(valid_files),
        "sheets_processed": len(sheet_order),
        "missing_files_count": 0,
        "folder_columns_count": 0,
        "camera_summary_found": 0,
        "camera_summary_missing": 0,
        "issues_count": 0,
        "output_path": output_path
    }

    for sname in sheet_order:
        ws_out = wb_out.create_sheet(title=sname)
        if hasattr(ws_out, 'views') and ws_out.views and ws_out.views.sheetView:
            ws_out.views.sheetView[0].showGridLines = True

        worksheets_for_this_sheet = []
        for fp, wb in loaded_wbs:
            if sname in wb.sheetnames:
                worksheets_for_this_sheet.append((fp, wb[sname]))

        if not worksheets_for_this_sheet:
            continue

        first_fp, first_ws = worksheets_for_this_sheet[0]

        # 1. Missing Files Sheet
        if sname.strip().lower() == 'missing files':
            for c in range(1, first_ws.max_column + 1):
                c_src = first_ws.cell(1, c)
                c_dst = ws_out.cell(1, c, value=c_src.value)
                copy_cell_style(c_src, c_dst)
                copy_col_dimensions(first_ws, ws_out, c, c)

            cur_row = 2
            for fp, ws in worksheets_for_this_sheet:
                for r in range(2, ws.max_row + 1):
                    row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                    if any(v is not None for v in row_vals):
                        for c in range(1, ws.max_column + 1):
                            src_c = ws.cell(r, c)
                            dst_c = ws_out.cell(cur_row, c, value=src_c.value)
                            copy_cell_style(src_c, dst_c)
                        cur_row += 1
                        stats["missing_files_count"] += 1

        # 2. Folder Structure Sheet
        elif sname.strip().lower() == 'folder structure':
            cur_dst_col = 1
            for fp, ws in worksheets_for_this_sheet:
                for c in range(1, ws.max_column + 1):
                    col_vals = [ws.cell(r, c).value for r in range(1, ws.max_row + 1)]
                    if any(v is not None for v in col_vals):
                        copy_col_dimensions(ws, ws_out, c, cur_dst_col)
                        for r in range(1, ws.max_row + 1):
                            src_c = ws.cell(r, c)
                            if src_c.value is not None or src_c.has_style:
                                dst_c = ws_out.cell(r, cur_dst_col, value=src_c.value)
                                copy_cell_style(src_c, dst_c)
                        cur_dst_col += 1
                        stats["folder_columns_count"] += 1

        # 3. Camera Summary Sheet
        elif sname.strip().lower() == 'camera summary':
            cameras = {}
            for fp, ws in worksheets_for_this_sheet:
                current_cam = None
                for r in range(1, ws.max_row + 1):
                    cell = ws.cell(r, 1)
                    val = str(cell.value or '').strip()
                    if not val:
                        continue

                    if re.match(r'^Cam\s*\d+', val, re.IGNORECASE) or val.startswith('Cam '):
                        current_cam = val
                        if current_cam not in cameras:
                            cameras[current_cam] = {
                                'items': [],
                                'total_found': 0,
                                'total_missing': 0,
                                'header_cell': cell,
                                'found_cell': None,
                                'missing_cell': None
                            }
                    elif val.startswith('Total Found:'):
                        m = re.search(r'Total Found:\s*(\d+)', val)
                        if m and current_cam:
                            cameras[current_cam]['total_found'] += int(m.group(1))
                            cameras[current_cam]['found_cell'] = cell
                    elif val.startswith('Total Missing:'):
                        m = re.search(r'Total Missing:\s*(\d+)', val)
                        if m and current_cam:
                            cameras[current_cam]['total_missing'] += int(m.group(1))
                            cameras[current_cam]['missing_cell'] = cell
                    else:
                        if current_cam:
                            cameras[current_cam]['items'].append((cell.value, cell))
                        else:
                            generic_cam = "Cam Summary"
                            if generic_cam not in cameras:
                                cameras[generic_cam] = {
                                    'items': [],
                                    'total_found': 0,
                                    'total_missing': 0,
                                    'header_cell': cell,
                                    'found_cell': None,
                                    'missing_cell': None
                                }
                            cameras[generic_cam]['items'].append((cell.value, cell))

            copy_col_dimensions(first_ws, ws_out, 1, 1)
            if ws_out.column_dimensions['A'].width is None:
                ws_out.column_dimensions['A'].width = 22.0

            cur_row = 1
            for cam_name, data in cameras.items():
                c_hdr = ws_out.cell(cur_row, 1, value=cam_name)
                copy_cell_style(data['header_cell'], c_hdr)
                cur_row += 1

                for val, src_cell in data['items']:
                    c_itm = ws_out.cell(cur_row, 1, value=val)
                    copy_cell_style(src_cell, c_itm)
                    cur_row += 1

                c_fnd = ws_out.cell(cur_row, 1, value=f"Total Found: {data['total_found']}")
                if data['found_cell']:
                    copy_cell_style(data['found_cell'], c_fnd)
                else:
                    c_fnd.font = Font(name="Segoe UI", size=11, bold=True)
                cur_row += 1

                c_mis = ws_out.cell(cur_row, 1, value=f"Total Missing: {data['total_missing']}")
                if data['missing_cell']:
                    copy_cell_style(data['missing_cell'], c_mis)
                    if data['total_missing'] == 0:
                        c_mis.fill = PatternFill(fill_type=None)
                else:
                    c_mis.font = Font(name="Segoe UI", size=11, bold=True)
                    if data['total_missing'] > 0:
                        c_mis.fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
                cur_row += 2

                stats["camera_summary_found"] += data['total_found']
                stats["camera_summary_missing"] += data['total_missing']

        # 4. Contributor, Issues, and other tabular sheets
        else:
            for c in range(1, first_ws.max_column + 1):
                c_src = first_ws.cell(1, c)
                c_dst = ws_out.cell(1, c, value=c_src.value)
                copy_cell_style(c_src, c_dst)
                copy_col_dimensions(first_ws, ws_out, c, c)

            cur_row = 2
            for fp, ws in worksheets_for_this_sheet:
                for r in range(2, ws.max_row + 1):
                    row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                    if any(v is not None for v in row_vals):
                        for c in range(1, ws.max_column + 1):
                            src_c = ws.cell(r, c)
                            dst_c = ws_out.cell(cur_row, c, value=src_c.value)
                            copy_cell_style(src_c, dst_c)
                        cur_row += 1
                        if sname.strip().lower() == 'issues':
                            stats["issues_count"] += 1

    wb_out.save(output_path)
    return stats


def merge_folder_reports(folder_path, output_filename=None):
    """Scans folder_path for all Excel inventory reports and merges them."""
    p = Path(folder_path)
    if not p.is_dir():
        raise NotADirectoryError(f"Directory not found: {folder_path}")

    xlsx_files = []
    for f in sorted(p.glob("*.xlsx")):
        if f.name.startswith("~$") or "MERGED" in f.name.upper():
            continue
        xlsx_files.append(str(f))

    if not xlsx_files:
        raise FileNotFoundError(f"No valid Excel reports found in {folder_path}")

    out_name = output_filename or "Combined_Merged_Inventory_Report.xlsx"
    out_path = str(p / out_name)
    return merge_inventory_reports(xlsx_files, out_path)


class ExcelMergerEngineWrapper(QObject):
    mergingChanged = Signal()
    mergeCompleted = Signal("QVariant")
    mergeFailed = Signal(str)
    scanCompleted = Signal("QVariant")

    def __init__(self, parent=None):
        super().__init__(parent)
        self._merging = False

    @Property(bool, notify=mergingChanged)
    def isMerging(self):
        return self._merging

    @Slot(str)
    def scanDirectory(self, folder_path):
        def _scan():
            try:
                p = Path(folder_path)
                if not p.is_dir():
                    self.scanCompleted.emit([])
                    return
                xlsx_files = []
                for f in sorted(p.glob("*.xlsx")):
                    if not f.name.startswith("~$") and "MERGED" not in f.name.upper():
                        xlsx_files.append(str(f))
                self.scanCompleted.emit(xlsx_files)
            except Exception as e:
                self.scanCompleted.emit([])
        threading.Thread(target=_scan, daemon=True).start()

    @Slot("QVariantList", str)
    def mergeFiles(self, file_paths, custom_out=""):
        if not file_paths:
            self.mergeFailed.emit("No Excel files selected for merging.")
            return

        self._merging = True
        self.mergingChanged.emit()

        def _worker():
            try:
                files = [str(f) for f in file_paths if f]
                out = custom_out.strip() if custom_out else None
                stats = merge_inventory_reports(files, out)
                self.mergeCompleted.emit(stats)
            except Exception as e:
                self.mergeFailed.emit(str(e))
            finally:
                self._merging = False
                self.mergingChanged.emit()

        threading.Thread(target=_worker, daemon=True).start()

    @Slot(str, str)
    def mergeFolder(self, folder_path, custom_out=""):
        if not folder_path or not os.path.isdir(folder_path):
            self.mergeFailed.emit("Invalid folder path.")
            return

        self._merging = True
        self.mergingChanged.emit()

        def _worker():
            try:
                stats = merge_folder_reports(folder_path, custom_out.strip() if custom_out else None)
                self.mergeCompleted.emit(stats)
            except Exception as e:
                self.mergeFailed.emit(str(e))
            finally:
                self._merging = False
                self.mergingChanged.emit()

        threading.Thread(target=_worker, daemon=True).start()


class AppReloaderWrapper(QObject):
    def __init__(self, app=None, parent=None):
        super().__init__(parent)
        self._app = app

    @Slot()
    def restartApp(self):
        """Reboots the entire application process seamlessly so any code/QML changes immediately reflect."""
        import subprocess
        python_exe = sys.executable
        main_script = os.path.abspath(sys.argv[0])
        try:
            subprocess.Popen([python_exe, main_script])
            if self._app:
                self._app.quit()
            else:
                sys.exit(0)
        except Exception as e:
            print(f"Error rebooting app: {e}")


class GoogleDriveManagerWrapper(QObject):
    configChanged = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._config_file = SCRIPT_DIR / "drive_config.json"
        self._config = {
            "mainDataUrl": "https://drive.google.com/drive/my-drive",
            "thumbnailsUrl": "https://drive.google.com/drive/my-drive"
        }
        self.load_config()

    def load_config(self):
        if self._config_file.exists():
            try:
                with open(self._config_file, "r", encoding="utf-8") as f:
                    self._config.update(json.load(f))
            except Exception:
                pass

    def save_config_file(self):
        try:
            with open(self._config_file, "w", encoding="utf-8") as f:
                json.dump(self._config, f, indent=2)
        except Exception:
            pass

    @Property(str, constant=True)
    def sessionStoragePath(self):
        storage_dir = os.path.abspath(str(SCRIPT_DIR / "drive_session_data")).replace("\\", "/")
        os.makedirs(storage_dir, exist_ok=True)
        return storage_dir

    @Property(str, constant=True)
    def mainSessionStoragePath(self):
        storage_dir = os.path.abspath(str(SCRIPT_DIR / "drive_session_data" / "main_drive")).replace("\\", "/")
        os.makedirs(storage_dir, exist_ok=True)
        return storage_dir

    @Property(str, constant=True)
    def refSessionStoragePath(self):
        storage_dir = os.path.abspath(str(SCRIPT_DIR / "drive_session_data" / "ref_drive")).replace("\\", "/")
        os.makedirs(storage_dir, exist_ok=True)
        return storage_dir

    @Property(str, constant=True)
    def defaultDownloadPath(self):
        downloads_dir = str(Path.home() / "Downloads").replace("\\", "/")
        os.makedirs(downloads_dir, exist_ok=True)
        return downloads_dir

    @Property(str, notify=configChanged)
    def mainDataUrl(self):
        return self._config.get("mainDataUrl", "https://drive.google.com/drive/my-drive")

    @Property(str, notify=configChanged)
    def thumbnailsUrl(self):
        return self._config.get("thumbnailsUrl", "https://drive.google.com/drive/my-drive")

    @Slot(str, str)
    def setDriveUrl(self, drive_type, url):
        clean_url = url.strip()
        if not clean_url.startswith("http"):
            clean_url = "https://" + clean_url
        if drive_type == "main":
            self._config["mainDataUrl"] = clean_url
        elif drive_type in ("thumbnails", "thumbs", "ref"):
            self._config["thumbnailsUrl"] = clean_url
        self.save_config_file()
        self.configChanged.emit()

    @Slot(str)
    def openInExternalBrowser(self, url):
        import webbrowser
        target = url.strip() if url else "https://drive.google.com"
        if not target.startswith("http"):
            target = "https://" + target
        webbrowser.open(target)

    @Slot(str)
    def clearSession(self, drive_type):
        """Clears local storage / session data for account re-authentication."""
        target_dir = self.mainSessionStoragePath if drive_type == "main" else self.refSessionStoragePath
        try:
            if os.path.exists(target_dir):
                shutil.rmtree(target_dir, ignore_errors=True)
                os.makedirs(target_dir, exist_ok=True)
        except Exception as e:
            print(f"Error clearing drive session ({drive_type}): {e}")



class ActivityAndReportsManagerWrapper(QObject):
    activityChanged = Signal()
    reportsChanged = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._activity_file = SCRIPT_DIR / "activity_log.json"
        self._reports_file = SCRIPT_DIR / "reports_log.json"
        self._activities = []
        self._reports = []
        self.load_data()

    def load_data(self):
        if self._activity_file.exists():
            try:
                with open(self._activity_file, "r", encoding="utf-8") as f:
                    self._activities = json.load(f)
            except Exception:
                self._activities = []
        if self._reports_file.exists():
            try:
                with open(self._reports_file, "r", encoding="utf-8") as f:
                    self._reports = json.load(f)
            except Exception:
                self._reports = []

    def save_data(self):
        try:
            with open(self._activity_file, "w", encoding="utf-8") as f:
                json.dump(self._activities[:100], f, indent=2)
            with open(self._reports_file, "w", encoding="utf-8") as f:
                json.dump(self._reports[:100], f, indent=2)
        except Exception:
            pass

    @Property("QVariantList", notify=activityChanged)
    def activities(self):
        return self._activities

    @Property("QVariantList", notify=reportsChanged)
    def reports(self):
        return self._reports

    @Slot(str, str, str, str, str)
    def logActivity(self, tool, action, details, color="#7C5CBF", bg="#F3EEFC"):
        entry = {
            "time": time.strftime("%I:%M %p"),
            "tool": tool,
            "action": action,
            "desc": details,
            "color": color,
            "bg": bg,
            "timestamp": time.time()
        }
        self._activities.insert(0, entry)
        self.save_data()
        self.activityChanged.emit()

    @Slot(str, str, str, str)
    def logReport(self, tool, file_count, result, path):
        entry = {
            "date": time.strftime("%d %b %Y"),
            "tool": tool,
            "files": str(file_count),
            "result": result,
            "path": path,
            "timestamp": time.time()
        }
        self._reports.insert(0, entry)
        self.save_data()
        self.reportsChanged.emit()


def main():
    SINGLE_INSTANCE_SOCKET = "SEQUORA_STUDIO_SINGLE_INSTANCE_V3"

    # 1. Single-Instance Check: If already running, activate existing window and exit
    client_socket = QLocalSocket()
    client_socket.connectToServer(SINGLE_INSTANCE_SOCKET)
    if client_socket.waitForConnected(300):
        try:
            client_socket.write(b"ACTIVATE\n")
            client_socket.waitForBytesWritten(300)
            client_socket.disconnectFromServer()
        except Exception:
            pass
        print("[NOTICE] SEQUORA Studio is already running. Switched focus to existing window.")
        sys.exit(0)

    # 2. Primary Instance Setup
    app = QApplication.instance() or QApplication(sys.argv)
    app.setApplicationName("SEQUORA Studio")
    app.setOrganizationName("BSFrameWorks")
    app.setOrganizationDomain("bsframeworks.com")

    # Set Application Icon
    icon_paths = [
        SCRIPT_DIR / "assets" / "icon.ico",
        SCRIPT_DIR / "assets" / "icon.png",
        PROJECT_ROOT / "assets" / "icon.ico",
        PROJECT_ROOT / "assets" / "icon.png"
    ]
    for ip in icon_paths:
        if ip.exists():
            app.setWindowIcon(QIcon(str(ip)))
            break


    engine = QQmlApplicationEngine()

    photo_engine = PhotoMatcherEngineWrapper()
    video_engine = VideoTransferEngineWrapper()
    pv_engine = PVSeparatorEngineWrapper()
    thumb_sep_engine = ThumbnailSeparatorEngineWrapper()
    remaining_engine = RemainingPhotosCollectorEngineWrapper()
    excel_merger_engine = ExcelMergerEngineWrapper()
    drive_engine = GoogleDriveManagerWrapper()
    activity_engine = ActivityAndReportsManagerWrapper()
    dialogs = NativeDialogsWrapper()
    reloader = AppReloaderWrapper(app)

    # Auto-wire backend signals to activity & report logging
    def on_photo_scan(res):
        st = res.get("stats", {}) if isinstance(res, dict) else {}
        activity_engine.logActivity(
            "Photo Matcher", "Scan completed",
            f"{st.get('total', 0)} photos scanned · {st.get('matched', 0)} matched",
            "#7C5CBF", "#F3EEFC"
        )
    photo_engine.scanCompleted.connect(on_photo_scan)

    def on_photo_rename(s, e, r_path, skipped=0):
        activity_engine.logActivity(
            "Photo Matcher", "Rename completed",
            f"{s} files tagged with _U/_R statuses",
            "#7C5CBF", "#F3EEFC"
        )
        if r_path:
            activity_engine.logReport("Photo Matcher", str(s), "Report exported to Documents folder", str(r_path))
    photo_engine.renameCompleted.connect(on_photo_rename)

    def on_video_transfer(summary, transferred_files):
        count = len(transferred_files) if isinstance(transferred_files, list) else 0
        activity_engine.logActivity(
            "Video Matcher", "Transfer completed",
            f"{count} video clips transferred successfully",
            "#059669", "#ECFDF5"
        )
    video_engine.transferCompleted.connect(on_video_transfer)

    def on_pv_extract(res):
        c = res.get("success_count", 0) if isinstance(res, dict) else 0
        dest = res.get("dest_dir", "") if isinstance(res, dict) else ""
        activity_engine.logActivity(
            "PV Separator", "Extraction completed",
            f"{c} photos extracted to Photo Data folder",
            "#D97706", "#FFFBEB"
        )
        if dest:
            activity_engine.logReport("PV Separator", str(c), "Photos separated and moved to destination", str(dest))
    pv_engine.processingCompleted.connect(on_pv_extract)

    def on_thumb_sep(res):
        c = res.get("success_count", 0) if isinstance(res, dict) else 0
        dest = res.get("dest_dir", "") if isinstance(res, dict) else ""
        activity_engine.logActivity(
            "Thumbnail Shifter", "Thumbnails separated",
            f"{c} thumbnails (_P/_V) organized by folder structure",
            "#06B6D4", "#F0F9FF"
        )
        if dest:
            activity_engine.logReport("Thumbnail Shifter", str(c), "Thumbnails shifted by folder structure", str(dest))
    thumb_sep_engine.processingCompleted.connect(on_thumb_sep)

    def on_remaining_shift(res):
        fc = res.get("shifted_folders", 0) if isinstance(res, dict) else 0
        activity_engine.logActivity(
            "Remaining Shifter", "Date shift completed",
            f"{fc} remaining folders consolidated",
            "#DB2777", "#FDF2F8"
        )
    remaining_engine.shiftCompleted.connect(on_remaining_shift)

    def on_excel_merge(res):
        fc = res.get("files_merged", 0) if isinstance(res, dict) else 0
        out_p = res.get("output_path", "") if isinstance(res, dict) else ""
        activity_engine.logActivity(
            "Report Merger", "Merge completed",
            f"{fc} inventory reports merged with full styles",
            "#10B981", "#ECFDF5"
        )
        if out_p:
            activity_engine.logReport("Report Merger", str(fc), "Excel workbook merged with styles", str(out_p))
    excel_merger_engine.mergeCompleted.connect(on_excel_merge)

    from updater import AppUpdaterWrapper
    app_updater = AppUpdaterWrapper()

    engine.rootContext().setContextProperty("photoEngine", photo_engine)
    engine.rootContext().setContextProperty("videoEngine", video_engine)
    engine.rootContext().setContextProperty("pvSeparatorEngine", pv_engine)
    engine.rootContext().setContextProperty("thumbSeparatorEngine", thumb_sep_engine)
    engine.rootContext().setContextProperty("thumbEngine", thumb_sep_engine)
    engine.rootContext().setContextProperty("remainingEngine", remaining_engine)
    engine.rootContext().setContextProperty("excelMergerEngine", excel_merger_engine)
    engine.rootContext().setContextProperty("driveEngine", drive_engine)
    engine.rootContext().setContextProperty("activityEngine", activity_engine)
    engine.rootContext().setContextProperty("nativeDialogs", dialogs)
    engine.rootContext().setContextProperty("appReloader", reloader)
    engine.rootContext().setContextProperty("appUpdater", app_updater)

    qml_file = SCRIPT_DIR / "qml" / "main.qml"
    engine.load(QUrl.fromLocalFile(str(SCRIPT_DIR / "qml" / "main.qml")))

    if not engine.rootObjects():
        print("[ERROR] Failed to load QML root object.")
        sys.exit(-1)

    # 3. Apply Crisp Native Windows Taskbar Icon via Win32 API
    def apply_native_taskbar_icon():
        ico_path = (SCRIPT_DIR / "assets" / "icon.ico").resolve()
        if not ico_path.exists():
            ico_path = (PROJECT_ROOT / "assets" / "icon.ico").resolve()
        if not ico_path.exists():
            return

        q_icon = QIcon(str(ico_path))
        for win in engine.rootObjects():
            if hasattr(win, "setIcon"):
                win.setIcon(q_icon)
            if sys.platform == "win32" and hasattr(win, "winId"):
                try:
                    import ctypes
                    WM_SETICON = 0x0080
                    ICON_SMALL = 0
                    ICON_BIG = 1
                    IMAGE_ICON = 1
                    LR_LOADFROMFILE = 0x00000010
                    LR_DEFAULTSIZE = 0x00000040

                    hicon_big = ctypes.windll.user32.LoadImageW(None, str(ico_path), IMAGE_ICON, 0, 0, LR_LOADFROMFILE | LR_DEFAULTSIZE)
                    hicon_small = ctypes.windll.user32.LoadImageW(None, str(ico_path), IMAGE_ICON, 16, 16, LR_LOADFROMFILE)
                    hwnd = int(win.winId())
                    if hwnd:
                        if hicon_big:
                            ctypes.windll.user32.SendMessageW(hwnd, WM_SETICON, ICON_BIG, hicon_big)
                        if hicon_small:
                            ctypes.windll.user32.SendMessageW(hwnd, WM_SETICON, ICON_SMALL, hicon_small)
                except Exception as e:
                    print("Taskbar icon attachment notice:", e)

    apply_native_taskbar_icon()

    # 4. Single-Instance Server: Listen for duplicate launch attempts and bring window to front
    local_server = QLocalServer()
    # Remove stale socket if needed
    QLocalServer.removeServer(SINGLE_INSTANCE_SOCKET)
    if local_server.listen(SINGLE_INSTANCE_SOCKET):
        def on_new_connection():
            sock = local_server.nextPendingConnection()
            if sock:
                sock.readyRead.connect(lambda: sock.readAll())
                # Restore and focus existing window
                for obj in engine.rootObjects():
                    if hasattr(obj, "showNormal"):
                        obj.showNormal()
                    if hasattr(obj, "raise_"):
                        obj.raise_()
                    if hasattr(obj, "requestActivate"):
                        obj.requestActivate()
                if sys.platform == "win32":
                    try:
                        import ctypes
                        hwnd = ctypes.windll.user32.FindWindowW(None, "SEQUORA Studio — Creative Production Suite")
                        if hwnd:
                            ctypes.windll.user32.ShowWindow(hwnd, 9)  # SW_RESTORE
                            ctypes.windll.user32.SetForegroundWindow(hwnd)
                    except Exception:
                        pass
        local_server.newConnection.connect(on_new_connection)

    print("[SUCCESS] SEQUORA Qt6/QML Native GUI launched.")
    sys.exit(app.exec())



if __name__ == "__main__":
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8")
        except Exception:
            pass
    if "--merge-folder" in sys.argv or "--folder" in sys.argv:
        idx = sys.argv.index("--merge-folder") if "--merge-folder" in sys.argv else sys.argv.index("--folder")
        folder = sys.argv[idx + 1] if idx + 1 < len(sys.argv) else r"C:\Users\Burhanuddin\Downloads\New folder"
        res = merge_folder_reports(folder)
        print("[SUCCESS] MERGED SUCCESSFULLY:", res["output_path"])
        sys.exit(0)
    elif "--merge-files" in sys.argv or "--files" in sys.argv:
        idx = sys.argv.index("--merge-files") if "--merge-files" in sys.argv else sys.argv.index("--files")
        files = sys.argv[idx + 1:]
        res = merge_inventory_reports(files)
        print("[SUCCESS] MERGED SUCCESSFULLY:", res["output_path"])
        sys.exit(0)
    else:
        main()
